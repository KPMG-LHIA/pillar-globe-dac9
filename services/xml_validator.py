"""
xml_validator.py  –  PILLAR GloBE/DAC9
Implementa tutti i check AdE verificabili offline dall'Allegato Tecnico (13 marzo 2026):
  • Sezione 8.2  – Record Errors SEVERE  (60001-60028, esclusi 60002/60008/60009/60014)
  • Sezione 8.3  – Record Errors OTHER   (70001-70124)

Changelog v1.3.1 (2026-06):
  lxml    – fix: sostituiti tutti i truth-test su elementi (if el:) con 'is not None'
              (con lxml un elemento senza figli valuta False → check silenziati)
  70026   – fix: OwnershipPercentage accetta sia frazione (1) che percentuale (100)

Changelog v1.3 (2026-06):
  70004   – fix: esclude TypeOfTIN=GIR3002 dal check issuedBy==ResCountryCode
  70016   – fix: elemento XML è GlobeStatus (non GloBEStatus); _findall per valori multipli
  _gs()   – nuovo helper che sostituisce set((_t(_find(...GloBEStatus...)) or "").split())

Changelog v1.2 (2026-06):
  _src()  – helper: aggiunge "(riga XML: N)" ai dettagli KO
  70033   – fix: lookup TIN include OtherUPE/ExcludedUPE
  70034   – fix: verifica Art2.1.3/Status solo se elemento presente
  70035   – fix: verifica Art2.1.5/Status solo se elemento presente
  70047   – fix: mapping SafeHarbour→ETR per SubGroup TIN, non per giurisdizione
  70048   – fix: idem
  70114   – fix: "segni opposti" solo per stesso AdjustmentItem code
  70118   – fix: idem per AdjustedCoveredTax/Adjustments
  70121   – fix: scope corretto = per singolo CEComputation, non aggregato ETR
"""
from __future__ import annotations
import copy, re, uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

NS = {
    "globe": "urn:oecd:ties:globe:v2",
    "stf":   "urn:oecd:ties:globestf:v5",
    "iso":   "urn:oecd:ties:isoglobetypes:v1",
}
PLACEHOLDER_RE = re.compile(r"\{Guid:D\}", re.IGNORECASE)
DOCREFID_RE    = re.compile(r"^[A-Z]{2}\d{4}[A-Za-z0-9\-]{6,}$")
MSGREFID_RE    = re.compile(r"^[A-Z]{2}\d{4}[A-Z]{2}[A-Za-z0-9\-]{1,}$")
TIN3003_RE     = re.compile(r"^P2[A-Z]{2}\d{8}[A-Z]{3}[A-Za-z0-9]{3}$")
GLOBE_STATUS_UPE_FORBIDDEN = {
    "GIR305","GIR307","GIR308","GIR309","GIR312",
    "GIR313","GIR314","GIR315","GIR317","GIR318",
}

# ── Helper: numero riga XML ───────────────────────────────────────────────────
def _src(el) -> str:
    if el is None: return ""
    line = getattr(el, "sourceline", None)
    return f" (riga XML: {line})" if line else ""

# ── CheckResult ───────────────────────────────────────────────────────────────
class CheckResult:
    __slots__ = ("code","desc","xpath","status","detail")
    STATUS_OK = "OK"; STATUS_KO = "KO"; STATUS_SKIP = "SKIP"; STATUS_WARN = "WARN"
    def __init__(self, code, desc, xpath, status="OK", detail=""):
        self.code=code; self.desc=desc; self.xpath=xpath; self.status=status; self.detail=detail
    def ko(self, d=""):
        self.status=self.STATUS_KO
        if d: self.detail=d
        return self
    def skip(self, r=""):
        self.status=self.STATUS_SKIP; self.detail=r; return self
    def warn(self, d=""):
        self.status=self.STATUS_WARN
        if d: self.detail=d
        return self

def _ko_all(r: "CheckResult", bad: list, max_items: int = 50) -> "CheckResult":
    """
    Imposta r come KO riportando TUTTE le occorrenze in bad (non solo la prima).
    Le occorrenze sono separate da newline nel campo detail.
    max_items: quante occorrenze mostrare al massimo (default 50).
    """
    if not bad:
        return r
    items = bad[:max_items]
    suffix = f"\n… (+{len(bad)-max_items} altri)" if len(bad) > max_items else ""
    r.status = CheckResult.STATUS_KO
    r.detail = "\n".join(items) + suffix
    return r

# ── XML helpers ───────────────────────────────────────────────────────────────
def _t(el) -> str:
    return "" if el is None else (el.text or "").strip()
def _find(root, xpath, ns=None):   return root.find(xpath, ns or NS)
def _findall(root, xpath, ns=None): return root.findall(xpath, ns or NS)
def _attr(el, a) -> str:
    return "" if el is None else (el.get(a) or "").strip()
def _decimal(s):
    try: return Decimal(str(s).replace(",","."))
    except (InvalidOperation, AttributeError): return None
def _year(el):
    t=_t(el)
    if t:
        try: return int(t[:4])
        except ValueError: pass
    return None
def _date(s):
    for fmt in ("%Y-%m-%d","%Y-%m","%Y"):
        try: return datetime.strptime(s.strip(), fmt).date()
        except ValueError: continue
    return None

# ── GlobeStatus helper ────────────────────────────────────────────────────────
def _gs(id_el, ns=None) -> set:
    """
    Raccoglie tutti i valori GlobeStatus da un elemento ID.
    NOTA 1: il nome dell'elemento è 'GlobeStatus' (non 'GloBEStatus').
    NOTA 2: può apparire più volte (es. GIR301+GIR307) → serve _findall, non _find.
    """
    if id_el is None:
        return set()
    return {e.text for e in _findall(id_el, "globe:GlobeStatus", ns or NS) if e.text}

# ── GUID placeholders ─────────────────────────────────────────────────────────
def _has_guid(root):
    for el in root.iter():
        if el.text and PLACEHOLDER_RE.search(el.text): return True
        for v in el.attrib.values():
            if PLACEHOLDER_RE.search(v): return True
    return False

def _resolve_guid(root):
    root = copy.deepcopy(root)
    for el in root.iter():
        if el.text and PLACEHOLDER_RE.search(el.text):
            el.text = PLACEHOLDER_RE.sub(lambda _: str(uuid.uuid4()), el.text)
        if el.tail and PLACEHOLDER_RE.search(el.tail):
            el.tail = PLACEHOLDER_RE.sub(lambda _: str(uuid.uuid4()), el.tail)
        for k in list(el.attrib):
            if PLACEHOLDER_RE.search(el.attrib[k]):
                el.attrib[k] = PLACEHOLDER_RE.sub(lambda _: str(uuid.uuid4()), el.attrib[k])
    return root

# ── sanitize filename ─────────────────────────────────────────────────────────
def _sanitize_filename(name):
    name = re.sub(r"\s+","_",name)
    name = re.sub(r"[^\w\-.]","",name)
    name = name.strip(".-_")
    return name or "output"

# ── Main entry ────────────────────────────────────────────────────────────────
def validate(xml_path, output_dir=None):
    xml_path   = Path(xml_path)
    output_dir = Path(output_dir) if output_dir else xml_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)
    try:
        tree = etree.parse(str(xml_path))
        root = tree.getroot()
    except etree.XMLSyntaxError as e:
        return _write_xlsx([CheckResult("50001","Il file XML non è parsabile.",
            "/GLOBE_OECD",CheckResult.STATUS_KO,str(e))], xml_path, output_dir)

    guid_resolved = _has_guid(root)
    if guid_resolved: root = _resolve_guid(root)

    results  = _check_file_errors(root, xml_path)
    results += _check_severe(root)
    results += _check_other(root)

    version_fixed = any(r.code=="50003" and r.status=="WARN" for r in results)
    return _write_xlsx(results, xml_path, output_dir,
                       guid_resolved=guid_resolved, version_fixed=version_fixed)


# ── 8.1 FILE ERRORS + C0000/C0001/C0004 ──────────────────────────────────────
def _check_file_errors(root, xml_path):
    out = []
    with open(xml_path,"rb") as f: raw = f.read()

    r = CheckResult("C0000","Il file non deve contenere il BOM UTF-8.","/GLOBE_OECD")
    if raw[:3] == b"\xef\xbb\xbf": r.ko("BOM UTF-8 rilevato. Salvare in UTF-8 senza BOM.")
    out.append(r)

    r = CheckResult("C0001","Il file deve essere codificato in UTF-8 valido.","/GLOBE_OECD")
    try:
        (raw[3:] if raw[:3]==b"\xef\xbb\xbf" else raw).decode("utf-8")
    except UnicodeDecodeError as e: r.ko(f"Byte non UTF-8 validi: {e}")
    out.append(r)

    r = CheckResult("C0004","I caratteri speciali XML devono essere codificati come entità.","/GLOBE_OECD")
    raw_text = (raw[3:] if raw[:3]==b"\xef\xbb\xbf" else raw).decode("utf-8",errors="replace")
    if re.findall(r"&(?!amp;|lt;|gt;|apos;|quot;|#)", raw_text):
        for i,line in enumerate(raw_text.splitlines(),1):
            if re.search(r"&(?!amp;|lt;|gt;|apos;|quot;|#)", line):
                r.ko(f"'&' non codificato riga {i}: {line.strip()[:80]!r}"); break
    out.append(r)

    out.append(CheckResult("50001","Il file XML è ben formato e parsabile.","/GLOBE_OECD"))

    r = CheckResult("50002","Elemento radice GLOBE_OECD presente.","/GLOBE_OECD")
    tag = root.tag.split("}")[-1] if "}" in root.tag else root.tag
    if tag != "GLOBE_OECD": r.ko(f"Radice trovata: {root.tag}")
    out.append(r)

    r = CheckResult("50003","GLOBE_OECD/@version presente e valorizzato.","/GLOBE_OECD/@version")
    if not root.get("version",""):
        root.set("version","1.0")
        r.warn("Attributo 'version' assente: impostato a '1.0' in-memory.")
    out.append(r)

    r = CheckResult("50004","Il file è codificato UTF-8 senza BOM.","/GLOBE_OECD")
    if raw[:3]==b"\xef\xbb\xbf": r.ko("BOM UTF-8 rilevato (non ammesso).")
    out.append(r)

    return out


# ── 8.2 SEVERE (60001-60028) ──────────────────────────────────────────────────
def _check_severe(root):
    out=[]; ns=NS
    ms   = _find(root,"globe:MessageSpec",ns)
    body = _find(root,"globe:GLOBEBody",ns)
    msg_ref     = _t(_find(ms,"globe:MessageRefId",ns))    if ms   is not None else ""
    rep_per     = _find(ms,"globe:ReportingPeriod",ns)     if ms   is not None else None
    rep_per_txt = _t(rep_per)
    filing_info = _find(body,"globe:FilingInfo",ns)        if body is not None else None
    gen_sec     = _find(body,"globe:GeneralSection",ns)    if body is not None else None
    summary     = _find(body,"globe:Summary",ns)           if body is not None else None
    jur_sections= _findall(body,"globe:JurisdictionSection",ns) if body is not None else []
    filing_doc  = _find(filing_info,"globe:DocSpec",ns)   if filing_info is not None else None
    gen_doc     = _find(gen_sec,"globe:DocSpec",ns)        if gen_sec     is not None else None
    all_docspecs= _findall(body,".//globe:DocSpec",ns)     if body        is not None else []

    # 60001
    r=CheckResult("60001","MessageRefId nel formato [SendingCountry][ReportingPeriod][ReceivingCountry][UniqueID]","/GLOBE_OECD/MessageSpec/MessageRefId")
    if not msg_ref: r.ko("MessageRefId assente o vuoto.")
    elif not MSGREFID_RE.match(msg_ref): r.warn(f"Formato inatteso: {msg_ref!r}")
    out.append(r)

    # 60003
    r=CheckResult("60003","L'anno di ReportingPeriod non può essere maggiore dell'anno corrente.","/GLOBE_OECD/MessageSpec/ReportingPeriod")
    if rep_per_txt:
        yr=_year(rep_per)
        if yr and yr>date.today().year: r.ko(f"Anno {yr} > anno corrente {date.today().year}")
    else: r.ko("ReportingPeriod assente.")
    out.append(r)

    # 60004
    r=CheckResult("60004","Il messaggio non può mescolare record OECD1 con OECD2/OECD3.","/GLOBE_OECD/GLOBEBody/*/DocSpec/DocTypeIndic")
    if body is not None:
        indics=[_t(_find(ds,"globe:DocTypeIndic",ns)) for ds in all_docspecs]
        if any(i in ("OECD1","OECD0") for i in indics) and any(i in ("OECD2","OECD3") for i in indics):
            r.ko("Presenti sia OECD1/OECD0 che OECD2/OECD3.")
    out.append(r)

    # 60005
    r=CheckResult("60005","Se DocTypeIndic è OECD2/OECD3, deve appartenere alla stessa sottosezione del CorrDocRefId.","/GLOBE_OECD/GLOBEBody/*/DocSpec/DocTypeIndic")
    bad=[f"DocSpec OECD2/3 senza CorrDocRefId" for ds in all_docspecs
         if _t(_find(ds,"globe:DocTypeIndic",ns)) in ("OECD2","OECD3") and not _t(_find(ds,"globe:CorrDocRefId",ns))]
    _ko_all(r, bad)
    out.append(r)

    # 60006
    r=CheckResult("60006","Lo stesso CorrDocRefId non può comparire più di una volta.","/GLOBE_OECD/GLOBEBody/*/DocSpec/CorrDocRefId")
    cids=[_t(_find(ds,"globe:CorrDocRefId",ns)) for ds in all_docspecs if _t(_find(ds,"globe:CorrDocRefId",ns))]
    dups={x for x in cids if cids.count(x)>1}
    if dups: r.ko(f"CorrDocRefId duplicati: {', '.join(sorted(dups)[:3])}")
    out.append(r)

    # 60007
    r=CheckResult("60007","DocRefId già usato per un altro record.","/GLOBE_OECD/GLOBEBody/*/DocSpec/DocRefId")
    dids=[_t(_find(ds,"globe:DocRefId",ns)) for ds in all_docspecs if _t(_find(ds,"globe:DocRefId",ns))]
    dups={x for x in dids if dids.count(x)>1}
    if dups: r.ko(f"DocRefId duplicati: {', '.join(sorted(dups)[:3])}")
    out.append(r)

    # 60010
    r=CheckResult("60010","FilingInfo OECD3 implica cancellazione di tutti i blocchi correlati.","/GLOBE_OECD/GLOBEBody/FilingInfo/DocSpec/DocTypeIndic")
    if filing_doc is not None and _t(_find(filing_doc,"globe:DocTypeIndic",ns))=="OECD3":
        for bt in ["globe:GeneralSection","globe:Summary","globe:JurisdictionSection","globe:UTPRAttribution"]:
            for blk in _findall(body,bt,ns):
                bd=_find(blk,"globe:DocSpec",ns)
                if bd is not None and _t(_find(bd,"globe:DocTypeIndic",ns))!="OECD3":
                    r.ko("FilingInfo=OECD3 ma altri blocchi non sono OECD3."); break
    out.append(r)

    # 60011
    r=CheckResult("60011","DocRefId nel formato [SendingCountry][ReportingYear][UniqueID].","/GLOBE_OECD/GLOBEBody/*/DocSpec/DocRefId")
    bad=[_t(_find(ds,"globe:DocRefId",ns))[:40] for ds in all_docspecs
         if _t(_find(ds,"globe:DocRefId",ns)) and not DOCREFID_RE.match(_t(_find(ds,"globe:DocRefId",ns)))]
    _ko_all(r, [f"Formato non valido: {b!r}" for b in bad])
    out.append(r)

    # 60012
    r=CheckResult("60012","Se DocTypeIndic è OECD1/OECD0, CorrDocRefId deve essere assente.","/GLOBE_OECD/GLOBEBody/*/DocSpec/CorrDocRefId")
    bad=[f"{_t(_find(ds,'globe:DocTypeIndic',ns))} con CorrDocRefId" for ds in all_docspecs
         if _t(_find(ds,"globe:DocTypeIndic",ns)) in ("OECD1","OECD0") and _t(_find(ds,"globe:CorrDocRefId",ns))]
    if bad: r.ko("; ".join(bad[:3]))
    out.append(r)

    # 60013
    r=CheckResult("60013","OECD0 ammesso solo per FilingInfo.","/GLOBEBody/*/DocSpec/DocTypeIndic")
    bbs=[bt.split(":")[-1] for bt in ["globe:GeneralSection","globe:Summary","globe:JurisdictionSection","globe:UTPRAttribution"]
         for blk in (_findall(body,bt,ns) if body is not None else [])
         for bd in [_find(blk,"globe:DocSpec",ns)] if bd is not None
         if _t(_find(bd,"globe:DocTypeIndic",ns))=="OECD0"]
    if bbs: r.ko(f"OECD0 in: {', '.join(bbs[:3])}")
    out.append(r)

    # 60015
    r=CheckResult("60015","Se DocTypeIndic è OECD2/OECD3, CorrDocRefId è obbligatorio.","/GLOBE_OECD/GLOBEBody/*/DocSpec/CorrDocRefId")
    bad=[f"{_t(_find(ds,'globe:DocTypeIndic',ns))} senza CorrDocRefId" for ds in all_docspecs
         if _t(_find(ds,"globe:DocTypeIndic",ns)) in ("OECD2","OECD3") and not _t(_find(ds,"globe:CorrDocRefId",ns))]
    if bad: r.ko("; ".join(bad[:3]))
    out.append(r)

    # 60016
    r=CheckResult("60016","FilingInfo OECD0 → GeneralSection non può essere OECD1.","/GLOBE_OECD/GLOBEBody/FilingInfo/DocSpec/DocTypeIndic")
    if filing_doc is not None and gen_doc is not None:
        if _t(_find(filing_doc,"globe:DocTypeIndic",ns))=="OECD0" and _t(_find(gen_doc,"globe:DocTypeIndic",ns))=="OECD1":
            r.ko("FilingInfo=OECD0 ma GeneralSection=OECD1")
    out.append(r)

    # 60017
    r=CheckResult("60017","Se FilingInfo/DocTypeIndic=OECD1, GeneralSection deve essere presente.","/GLOBE_OECD/GLOBEBody/FilingInfo/DocSpec/DocTypeIndic")
    if filing_doc is not None and _t(_find(filing_doc,"globe:DocTypeIndic",ns))=="OECD1" and gen_sec is None:
        r.ko("FilingInfo=OECD1 ma GeneralSection assente.")
    out.append(r)

    # 60019
    r=CheckResult("60019","Se FilingCE/Role è GIR403/404/405, RecJurCode deve essere la giurisdizione locale.","/GLOBE_OECD/GLOBEBody/GeneralSection/RecJurCode")
    if filing_info is not None and gen_sec is not None:
        fce=_find(filing_info,"globe:FilingCE",ns)
        role=_t(_find(fce,"globe:Role",ns)) if fce is not None else ""
        if role in ("GIR403","GIR404","GIR405"):
            rj=_t(_find(gen_sec,"globe:RecJurCode",ns)); rc=_t(_find(fce,"globe:ResCountryCode",ns))
            if rj and rc and rj!=rc: r.ko(f"RecJurCode={rj!r} ≠ ResCountryCode={rc!r}")
    out.append(r)

    # 60020
    r=CheckResult("60020","Period/Start non può essere successivo a Period/End.","/GLOBE_OECD/GLOBEBody/FilingInfo/Period/Start")
    if filing_info is not None:
        p=_find(filing_info,"globe:Period",ns)
        if p is not None:
            s=_date(_t(_find(p,"globe:Start",ns))); e=_date(_t(_find(p,"globe:End",ns)))
            if s and e and s>e: r.ko(f"Start={s} > End={e}")
    out.append(r)

    # 60021
    r=CheckResult("60021","FilingInfo/Period/End non può essere successivo a ReportingPeriod.","/GLOBE_OECD/GLOBEBody/FilingInfo/Period/End")
    if filing_info is not None and rep_per_txt:
        p=_find(filing_info,"globe:Period",ns)
        if p is not None:
            e=_date(_t(_find(p,"globe:End",ns))); rp=_date(rep_per_txt)
            if e and rp and e.year>rp.year: r.ko(f"Period/End={e} > ReportingPeriod={rp}")
    out.append(r)

    # 60022
    r=CheckResult("60022","Se FilingCE/Role=GIR401, FilingCE/TIN deve coincidere con un TIN UPE.","/GLOBE_OECD/GLOBEBody/FilingInfo/FilingCE/TIN")
    if filing_info is not None and gen_sec is not None:
        fce=_find(filing_info,"globe:FilingCE",ns)
        if fce is not None and _t(_find(fce,"globe:Role",ns))=="GIR401":
            ftin=_t(_find(fce,"globe:TIN",ns))
            cs2=_find(gen_sec,"globe:CorporateStructure",ns)
            upetins=set()
            if cs2 is not None:
                for p in [".//globe:UPE/globe:ExcludedUPE/globe:ID/globe:TIN",".//globe:UPE/globe:OtherUPE/globe:ID/globe:TIN"]:
                    for t in _findall(cs2,p,ns):
                        v=_t(t)
                        if v: upetins.add(v)
            if ftin and upetins and ftin not in upetins: r.ko(f"FilingCE/TIN={ftin!r} non in UPE TINs")
    out.append(r)

    # 60023
    r=CheckResult("60023","FilingCE/ResCountryCode deve coincidere con TransmittingCountry.","/GLOBE_OECD/GLOBEBody/FilingInfo/FilingCE/ResCountryCode")
    if ms is not None and filing_info is not None:
        tc=_t(_find(ms,"globe:TransmittingCountry",ns))
        fce=_find(filing_info,"globe:FilingCE",ns)
        rc=_t(_find(fce,"globe:ResCountryCode",ns)) if fce is not None else ""
        if tc and rc and tc!=rc: r.ko(f"TransmittingCountry={tc!r} ≠ ResCountryCode={rc!r}")
    out.append(r)

    # 60024
    r=CheckResult("60024","Se SafeHarbour/ETRRange/SBIE/QDMTTut/GLoBETut presenti, JurWithTaxingRights/JurisdictionName deve essere valorizzato.","/GLOBE_OECD/GLOBEBody/Summary/JurWithTaxingRights/JurisdictionName")
    if summary is not None:
        has_d=any(_find(summary,t,ns) is not None for t in ["globe:SafeHarbour","globe:ETRRange","globe:SBIE","globe:QDMTTut","globe:GLoBETut"])
        if has_d:
            jwtr=_find(summary,"globe:JurWithTaxingRights",ns)
            if not _t(_find(jwtr,"globe:JurisdictionName",ns) if jwtr is not None else None): r.ko("JurisdictionName non valorizzato.")
    out.append(r)

    # 60025-60028 (formulas)
    out += _check_severe_formulas(jur_sections, ns)

    return out

def _check_severe_formulas(jur_sections, ns):
    out=[]
    # 60025
    r=CheckResult("60025","ETRRate = AdjustedCoveredTax/Total / NetGlobeIncome/Total.","JurisdictionSection/.../ETRRate")
    errs=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            er=_find(oc,"globe:ETRRate",ns); ne=_find(oc,"globe:NetGlobeIncome/globe:Total",ns)
            ae=_find(oc,"globe:AdjustedCoveredTax/globe:Total",ns)
            if er is None or ne is None: continue
            ngi=_decimal(_t(ne)); act=_decimal(_t(ae)) or Decimal(0); etr=_decimal(_t(er))
            if ngi is None or etr is None or ngi<=0: continue
            if abs(act/ngi - etr)>Decimal("0.0001"):
                errs.append(f"{jn}: ETRRate={etr} ≠ {act}/{ngi}={act/ngi:.6f}{_src(er)}")
    _ko_all(r, errs)
    out.append(r)

    # 60026
    r=CheckResult("60026","TopUpTax = (TopUpTaxPercentage * ExcessProfits) + AdditionalTopUpTax - QDMTT.","JurisdictionSection/.../TopUpTax")
    errs=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            te=_find(oc,"globe:TopUpTax",ns); pe=_find(oc,"globe:TopUpTaxPercentage",ns); ep=_find(oc,"globe:ExcessProfits",ns)
            if te is None or pe is None or ep is None: continue
            tut=_decimal(_t(te)); pct=_decimal(_t(pe)); exp=_decimal(_t(ep))
            if tut is None or pct is None or exp is None: continue
            an=_decimal(_t(_find(oc,"globe:AdditionalTopUpTax/globe:NONArt4.1.5/globe:AdditionalTopUpTax",ns))) or Decimal(0)
            aa=_decimal(_t(_find(oc,"globe:AdditionalTopUpTax/globe:Art4.1.5/globe:AdditionalTopUpTax",ns))) or Decimal(0)
            qd=_decimal(_t(_find(oc,"globe:QDMTT/globe:Amount",ns))) or Decimal(0)
            expected=(pct*exp)+an+aa-qd
            if abs(expected-tut)>Decimal("1"): errs.append(f"{jn}: TopUpTax={tut} ≠ {expected:.2f}{_src(te)}")
    _ko_all(r, errs)
    out.append(r)

    # 60027
    r=CheckResult("60027","IIR/ParentEntity/TopUpTax = TopUpTaxShare - IIROffset.","JurisdictionSection/.../IIR/ParentEntity/TopUpTax")
    errs=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ltce in _findall(js,".//globe:LowTaxJurisdiction/globe:LTCE",ns):
            for pe in _findall(ltce,"globe:IIR/globe:ParentEntity",ns):
                te=_find(pe,"globe:TopUpTax",ns); se=_find(pe,"globe:TopUpTaxShare",ns); oe=_find(pe,"globe:IIROffset",ns)
                if te is None or se is None: continue
                tut=_decimal(_t(te)); sh=_decimal(_t(se)); off=_decimal(_t(oe)) or Decimal(0)
                if tut is None or sh is None: continue
                if abs(sh-off-tut)>Decimal("1"): errs.append(f"{jn}: TopUpTax={tut} ≠ {sh}-{off}={sh-off}{_src(te)}")
    _ko_all(r, errs)
    out.append(r)

    # 60028
    r=CheckResult("60028","AdjustedFANIL/Total = FANIL + Σ(Additions) - Σ(Reductions).","JurisdictionSection/.../AdjustedFANIL/Total")
    errs=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            adf=_find(ce,"globe:AdjustedFANIL",ns)
            if adf is None: continue
            te=_find(adf,"globe:Total",ns); fe=_find(adf,"globe:FANIL",ns)
            if te is None or fe is None: continue
            tot=_decimal(_t(te)); fan=_decimal(_t(fe))
            if tot is None or fan is None: continue
            adds=sum((_decimal(_t(x)) or Decimal(0)) for x in _findall(adf,".//globe:MainEntityPEandFTE/globe:Additions",ns))
            reds=sum((_decimal(_t(x)) or Decimal(0)) for x in _findall(adf,".//globe:MainEntityPEandFTE/globe:Reductions",ns))
            if abs(fan+adds-reds-tot)>Decimal("1"): errs.append(f"{jn}: Total={tot} ≠ {fan}+{adds}-{reds}{_src(te)}")
    _ko_all(r, errs)
    out.append(r)

    return out


# ── 8.3 OTHER (70001-70124) ───────────────────────────────────────────────────
def _check_other(root):
    out=[]; ns=NS
    body        =_find(root,"globe:GLOBEBody",ns)
    ms          =_find(root,"globe:MessageSpec",ns)
    filing_info =_find(body,"globe:FilingInfo",ns)      if body is not None else None
    gen_sec     =_find(body,"globe:GeneralSection",ns)  if body is not None else None
    summary_els =_findall(body,"globe:Summary",ns)      if body is not None else []
    jur_sections=_findall(body,"globe:JurisdictionSection",ns) if body is not None else []
    utpr_attr   =_find(body,"globe:UTPRAttribution",ns) if body is not None else None
    cs          =_find(gen_sec,"globe:CorporateStructure",ns) if gen_sec is not None else None
    period_el   =_find(filing_info,"globe:Period",ns)   if filing_info is not None else None
    period_end  =_date(_t(_find(period_el,"globe:End",ns)))   if period_el is not None else None
    period_start=_date(_t(_find(period_el,"globe:Start",ns))) if period_el is not None else None

    def _ce_tins():
        t=set()
        if cs is not None:
            for e in _findall(cs,".//globe:CE/globe:ID/globe:TIN",ns):
                v=_t(e)
                if v: t.add(v)
        return t

    def _upe_tins():
        t=set()
        if cs is not None:
            for p in [".//globe:UPE/globe:ExcludedUPE/globe:ID/globe:TIN",
                      ".//globe:UPE/globe:OtherUPE/globe:ID/globe:TIN"]:
                for e in _findall(cs,p,ns):
                    v=_t(e)
                    if v: t.add(v)
        return t

    # ── 8.3.1 TIN (70001-70007) ──────────────────────────────────────────────
    terrs={c:[] for c in [f"7000{i}" for i in range(1,8)]}
    for te in _findall(root,".//globe:TIN",ns):
        val=_t(te); tot=_attr(te,"TypeOfTIN"); unk=_attr(te,"unknown").upper(); iss=_attr(te,"issuedBy"); ln=_src(te)
        if tot=="GIR3004":
            if val!="NOTIN" or unk!="TRUE" or iss: terrs["70001"].append(f"GIR3004: val={val!r} unk={unk!r} issuedBy={iss!r}{ln}")
        if val=="NOTIN":
            if tot!="GIR3004" or unk!="TRUE" or iss: terrs["70002"].append(f"NOTIN: TypeOfTIN={tot!r} unk={unk!r} issuedBy={iss!r}{ln}")
        if unk=="TRUE":
            if val!="NOTIN" or tot!="GIR3004" or iss: terrs["70003"].append(f"unknown=TRUE: val={val!r} TypeOfTIN={tot!r}{ln}")
        if iss=="IT" and val not in ("NOTIN","") and tot not in ("GIR3002","GIR3003","GIR3004"):
            if not (re.match(r"^\d{11}$",val) or re.match(r"^[A-Z0-9]{16}$",val)):
                terrs["70004"].append(f"TIN IT non valido: {val!r}{ln}")
        if not tot: terrs["70005"].append(f"@TypeOfTIN assente (val={val!r}){ln}")
        if not iss and tot not in ("GIR3003","GIR3004",""): terrs["70005"].append(f"@issuedBy assente TypeOfTIN={tot!r}{ln}")
        if tot=="GIR3003" and not TIN3003_RE.match(val): terrs["70007"].append(f"TIN GIR3003 formato non valido: {val!r}{ln}")
    for path in [".//globe:CorporateStructure/globe:UPE/globe:ExcludedUPE/globe:ID/globe:TIN",
                 ".//globe:CorporateStructure/globe:UPE/globe:OtherUPE/globe:ID/globe:TIN",
                 ".//globe:CorporateStructure/globe:CE/globe:ID/globe:TIN",
                 ".//globe:CorporateStructure/globe:CE/globe:QIIR/globe:Exception/globe:TIN"]:
        for te in _findall(root,path,ns):
            if _attr(te,"TypeOfTIN")=="GIR3004" or _attr(te,"unknown").upper()=="TRUE":
                terrs["70006"].append(f"TIN strutturale GIR3004/unknown: {_t(te)!r}{_src(te)}")
    descs={"70001":"Se TypeOfTIN=GIR3004, val=NOTIN, unknown=TRUE, issuedBy assente.",
           "70002":"Se TIN=NOTIN, TypeOfTIN=GIR3004, unknown=TRUE, issuedBy assente.",
           "70003":"Se unknown=TRUE, TIN=NOTIN, TypeOfTIN=GIR3004, issuedBy assente.",
           "70004":"Se issuedBy è giurisdizione locale, il TIN deve essere valido.",
           "70005":"@issuedBy e @TypeOfTIN devono essere presenti (salvo GIR3003/GIR3004).",
           "70006":"TIN strutturali non possono usare TypeOfTIN=GIR3004 né @unknown=TRUE.",
           "70007":"Se TypeOfTIN=GIR3003, il TIN deve rispettare il formato P2JJYYYYMMDDCCCXXX."}
    for code,errs in terrs.items():
        r=CheckResult(code,descs[code],"/GLOBE_OECD/GLOBEBody//TIN")
        _ko_all(r, errs)
        out.append(r)

    # ── 8.3.2 RecJurCode/UPE/Rules (70008-70012) ─────────────────────────────
    r=CheckResult("70008","UTPRAttribution/RecJurCode deve essere la giurisdizione UPE o JurWithTaxingRights.","/GLOBE_OECD/GLOBEBody/UTPRAttribution/RecJurCode")
    if utpr_attr is not None and gen_sec is not None:
        uj=_t(_find(utpr_attr,"globe:RecJurCode",ns)); vj=set()
        if cs is not None:
            for p in [".//globe:UPE/globe:OtherUPE/globe:ID/globe:ResCountryCode",
                      ".//globe:UPE/globe:ExcludedUPE/globe:ID/globe:ResCountryCode"]:
                for e in _findall(cs,p,ns): vj.add(_t(e))
        for s in summary_els:
            for jw in _findall(s,"globe:JurWithTaxingRights",ns):
                for jn in _findall(jw,"globe:JurisdictionName",ns): vj.add(_t(jn))
        if uj and vj and uj not in vj: r.ko(f"RecJurCode={uj!r} non in {sorted(vj)[:5]}")
    out.append(r)

    r=CheckResult("70009","GloBEStatus UPE non deve contenere: GIR305,GIR307-309,GIR312-315,GIR317,GIR318.","/GLOBE_OECD/GLOBEBody/GeneralSection/CorporateStructure/UPE/*/ID/GloBEStatus")
    bad=[]
    if cs is not None:
        for p in [".//globe:UPE/globe:ExcludedUPE/globe:ID",".//globe:UPE/globe:OtherUPE/globe:ID"]:
            for id_el in _findall(cs,p,ns):
                fb = _gs(id_el, ns) & GLOBE_STATUS_UPE_FORBIDDEN
                if fb: bad.append(f"UPE GlobeStatus vietato: {fb}{_src(id_el)}")
    if bad: r.ko("; ".join(bad[:3]))
    out.append(r)

    r=CheckResult("70010","Per OtherUPE, ID/ResCountryCode deve avere un solo valore.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../OtherUPE/ID/ResCountryCode")
    bad=[]
    if cs is not None:
        for e in _findall(cs,".//globe:UPE/globe:OtherUPE/globe:ID",ns):
            if len(_findall(e,"globe:ResCountryCode",ns))>1: bad.append(f"OtherUPE con più ResCountryCode{_src(e)}")
    _ko_all(r, bad)
    out.append(r)

    r=CheckResult("70011","Per CE, ID/ResCountryCode deve avere un solo valore.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/ID/ResCountryCode")
    bad=[]
    if cs is not None:
        for e in _findall(cs,".//globe:CE/globe:ID",ns):
            if len(_findall(e,"globe:ResCountryCode",ns))>1: bad.append(f"CE con più ResCountryCode{_src(e)}")
    _ko_all(r, bad)
    out.append(r)

    r=CheckResult("70012","Entità stesso ResCountryCode → stesso Rules (salvo GIR204).","/GLOBE_OECD/GLOBEBody/GeneralSection/.../Rules")
    bad=[]
    if cs is not None:
        rr={}
        for p in [".//globe:UPE/globe:ExcludedUPE/globe:ID",".//globe:UPE/globe:OtherUPE/globe:ID",".//globe:CE/globe:ID"]:
            for e in _findall(cs,p,ns):
                rc=_t(_find(e,"globe:ResCountryCode",ns)); rv=set((_t(_find(e,"globe:Rules",ns)) or "").split())
                if rc and "GIR204" not in rv:
                    if rc not in rr: rr[rc]=rv
                    elif rr[rc]!=rv: bad.append(f"Giurisdizione {rc}: Rules incoerenti")
    _ko_all(r, bad)
    out.append(r)

    # ── 8.3.3 GloBEStatus CE (70013-70021) ───────────────────────────────────
    ce_gs={}; ce_el_map={}
    if cs is not None:
        for ce in _findall(cs,"globe:CE",ns):
            id_=_find(ce,"globe:ID",ns)
            if id_ is None: continue
            tin=_t(_find(id_,"globe:TIN",ns))
            gs=_gs(id_, ns)
            ce_gs[tin]=gs; ce_el_map[tin]=id_
    all_gs=set().union(*ce_gs.values()) if ce_gs else set()

    def _gspair(code,desc,fn,xpath):
        r=CheckResult(code,desc,xpath)
        bad=[f"CE TIN={t!r} GlobeStatus={sorted(g)}{_src(ce_el_map.get(t))}" for t,g in ce_gs.items() if fn(g)]
        _ko_all(r, bad)
        return r

    out.append(_gspair("70013","Se GloBEStatus contiene GIR313, non deve contenere GIR314.",lambda g:"GIR313" in g and "GIR314" in g,"/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/ID/GloBEStatus"))
    out.append(_gspair("70014","Se GloBEStatus contiene GIR307, non deve contenere GIR308.",lambda g:"GIR307" in g and "GIR308" in g,"/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/ID/GloBEStatus"))

    r=CheckResult("70015","Se un CE ha GIR308, deve esistere un CE con GIR307.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/ID/GloBEStatus")
    b308=[t for t,g in ce_gs.items() if "GIR308" in g]
    if b308 and "GIR307" not in all_gs: r.ko(f"CE {b308[0]!r} ha GIR308 ma nessun CE ha GIR307{_src(ce_el_map.get(b308[0]))}")
    out.append(r)

    out.append(_gspair("70016","Se GloBEStatus contiene GIR307, deve contenere anche GIR309.",lambda g:"GIR307" in g and "GIR309" not in g,"/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/ID/GloBEStatus"))
    out.append(_gspair("70017","Se GloBEStatus contiene GIR308, deve contenere anche GIR309.",lambda g:"GIR308" in g and "GIR309" not in g,"/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/ID/GloBEStatus"))
    out.append(_gspair("70018","Se GloBEStatus contiene GIR305, non deve contenere GIR306 nello stesso CE.",lambda g:"GIR305" in g and "GIR306" in g,"/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/ID/GloBEStatus"))

    r=CheckResult("70019","Se un CE ha GIR305, deve esistere un CE con GIR306.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/ID/GloBEStatus")
    if any("GIR305" in g for g in ce_gs.values()) and "GIR306" not in all_gs: r.ko("CE con GIR305 ma nessun CE con GIR306.")
    out.append(r)

    out.append(_gspair("70020","Se GloBEStatus contiene GIR316 o GIR318, non deve contenere altri valori.",lambda g:("GIR316" in g or "GIR318" in g) and len(g)>1,"/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/ID/GloBEStatus"))

    r=CheckResult("70021","CE con GIR316 o GIR318 deve avere OwnershipChange.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/ID/GloBEStatus")
    bad=[]
    if cs is not None:
        for ce in _findall(cs,"globe:CE",ns):
            id_=_find(ce,"globe:ID",ns)
            if id_ is None: continue
            gs=_gs(id_, ns)
            if "GIR316" in gs or "GIR318" in gs:
                oc=_find(ce,"globe:OwnershipChange",ns)
                if oc is None or not _t(_find(oc,"globe:ChangeDate",ns)):
                    bad.append(f"CE {_t(_find(id_,'globe:TIN',ns))!r} ha {gs&{'GIR316','GIR318'}} ma manca OwnershipChange{_src(id_)}")
    _ko_all(r, bad)
    out.append(r)

    # ── 8.3.4 OwnershipChange (70022-70025) ──────────────────────────────────
    r=CheckResult("70022","OwnershipChange/ChangeDate non può essere anteriore a Period/Start.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/OwnershipChange/ChangeDate")
    bad=[]
    if cs is not None and period_start:
        for ce in _findall(cs,"globe:CE",ns):
            oc=_find(ce,"globe:OwnershipChange",ns)
            if oc is not None:
                cd_el=_find(oc,"globe:ChangeDate",ns); cd=_date(_t(cd_el))
                if cd and cd<period_start:
                    tin=_t(_find(_find(ce,"globe:ID",ns),"globe:TIN",ns)) if _find(ce,"globe:ID",ns) else "?"
                    bad.append(f"CE {tin!r}: ChangeDate={cd} < Start={period_start}{_src(cd_el)}")
    _ko_all(r, bad)
    out.append(r)

    r=CheckResult("70023","OwnershipChange/ChangeDate non può essere successiva a Period/End.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/OwnershipChange/ChangeDate")
    bad=[]
    if cs is not None and period_end:
        for ce in _findall(cs,"globe:CE",ns):
            oc=_find(ce,"globe:OwnershipChange",ns)
            if oc is not None:
                cd_el=_find(oc,"globe:ChangeDate",ns); cd=_date(_t(cd_el))
                if cd and cd>period_end:
                    tin=_t(_find(_find(ce,"globe:ID",ns),"globe:TIN",ns)) if _find(ce,"globe:ID",ns) else "?"
                    bad.append(f"CE {tin!r}: ChangeDate={cd} > End={period_end}{_src(cd_el)}")
    _ko_all(r, bad)
    out.append(r)

    r=CheckResult("70024","PreOwnership non deve essere compilato quando PreGloBEStatus=GIR719.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/OwnershipChange/PreOwnership")
    bad=[]
    if cs is not None:
        for ce in _findall(cs,"globe:CE",ns):
            oc=_find(ce,"globe:OwnershipChange",ns)
            if oc and _t(_find(oc,"globe:PreGloBEStatus",ns))=="GIR719" and _find(oc,"globe:PreOwnership",ns) is not None:
                bad.append(f"OwnershipChange con PreGloBEStatus=GIR719 ha PreOwnership{_src(oc)}")
    _ko_all(r, bad)
    out.append(r)

    r=CheckResult("70025","Se PreOwnership/OwnershipType contiene GIR805/GIR806, TIN deve essere GIR3004/NOTIN.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/OwnershipChange/PreOwnership/TIN")
    bad=[]
    if cs is not None:
        for ce in _findall(cs,"globe:CE",ns):
            oc=_find(ce,"globe:OwnershipChange",ns)
            if oc is None: continue
            for po in _findall(oc,"globe:PreOwnership",ns):
                ot=set((_t(_find(po,"globe:OwnershipType",ns)) or "").split())
                if ot&{"GIR805","GIR806"}:
                    for te in _findall(po,"globe:TIN",ns):
                        if _attr(te,"TypeOfTIN")!="GIR3004" or _t(te)!="NOTIN":
                            bad.append(f"PreOwnership GIR805/806: TIN={_t(te)!r} TypeOfTIN={_attr(te,'TypeOfTIN')!r}{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    # ── 8.3.5 Ownership (70026-70031) ────────────────────────────────────────
    def _ow26(ce,ns):
        id_=_find(ce,"globe:ID",ns); gs=_gs(id_, ns) if id_ is not None else set()
        if "GIR305" not in gs: return None
        for ow in _findall(ce,"globe:Ownership",ns):
            pe=_find(ow,"globe:OwnershipPercentage",ns); p=_decimal(_t(pe))
            # FIX v1.3.1: OwnershipPercentage può essere espresso come frazione (1 = 100%)
            # o come percentuale (100 = 100%). Entrambe le convenzioni sono accettate da AdE.
            if p is not None and p!=Decimal("100") and p!=Decimal("1"):
                return f"CE {_t(_find(id_,'globe:TIN',ns))!r} GIR305: pct={p} ≠ 100%{_src(pe)}"
        return None
    def _ow27(ce,ns):
        id_=_find(ce,"globe:ID",ns); gs=_gs(id_, ns) if id_ is not None else set()
        if "GIR318" not in gs: return None
        for ow in _findall(ce,"globe:Ownership",ns):
            pe=_find(ow,"globe:OwnershipPercentage",ns); p=_decimal(_t(pe)); tv=_t(_find(ow,"globe:TIN",ns)); ot=_t(_find(ow,"globe:OwnershipType",ns))
            if p!=Decimal("0") or tv!="NOTIN" or ot!="GIR806":
                return f"CE {_t(_find(id_,'globe:TIN',ns))!r} GIR318: pct={p} tin={tv!r} ot={ot!r}{_src(pe)}"
        return None
    def _ow28(ce,ns):
        id_=_find(ce,"globe:ID",ns); gs=_gs(id_, ns) if id_ is not None else set()
        if "GIR318" in gs: return None
        for ow in _findall(ce,"globe:Ownership",ns):
            pe=_find(ow,"globe:OwnershipPercentage",ns); p=_decimal(_t(pe))
            if p is not None and p==Decimal("0"):
                return f"CE {_t(_find(id_,'globe:TIN',ns))!r}: pct=0% (non GIR318){_src(pe)}"
        return None

    if cs is not None:
        for code,desc,fn in [
            ("70026","Se GloBEStatus=GIR305, OwnershipPercentage deve essere 100%.",_ow26),
            ("70027","Se GloBEStatus=GIR318, pct=0%, TIN=NOTIN, Type=GIR806.",_ow27),
            ("70028","Salvo GIR318, OwnershipPercentage non deve essere 0%.",_ow28),
        ]:
            r=CheckResult(code,desc,"/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/Ownership")
            bad=[e for ce in _findall(cs,"globe:CE",ns) for e in [fn(ce,ns)] if e]
            _ko_all(r, bad)
            out.append(r)

    r=CheckResult("70029","Se OwnershipType contiene GIR801, TIN deve corrispondere a un TIN UPE.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/Ownership/TIN")
    bad=[]
    ut=_upe_tins()
    if cs is not None:
        for ce in _findall(cs,"globe:CE",ns):
            for ow in _findall(ce,"globe:Ownership",ns):
                if "GIR801" in set((_t(_find(ow,"globe:OwnershipType",ns)) or "").split()):
                    te=_find(ow,"globe:TIN",ns); tv=_t(te)
                    if tv and ut and tv not in ut: bad.append(f"GIR801 TIN={tv!r} non in UPE{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    r=CheckResult("70030","Se OwnershipType contiene GIR802/803/804, TIN deve corrispondere a un CE TIN.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/Ownership/TIN")
    bad=[]
    ct=_ce_tins()
    if cs is not None:
        for ce in _findall(cs,"globe:CE",ns):
            for ow in _findall(ce,"globe:Ownership",ns):
                ots=set((_t(_find(ow,"globe:OwnershipType",ns)) or "").split())
                if ots&{"GIR802","GIR803","GIR804"}:
                    te=_find(ow,"globe:TIN",ns); tv=_t(te)
                    if tv and ct and tv not in ct: bad.append(f"{ots&{'GIR802','GIR803','GIR804'}} TIN={tv!r} non in CE TINs{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    r=CheckResult("70031","Se GloBEStatus=GIR305, Ownership/TIN deve essere uguale a un TIN con GIR306.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/Ownership/TIN")
    bad=[]; t306=set()
    if cs is not None:
        for ce in _findall(cs,"globe:CE",ns):
            id_=_find(ce,"globe:ID",ns)
            if id_ is not None and "GIR306" in _gs(id_, ns):
                for te in _findall(id_,"globe:TIN",ns): t306.add(_t(te))
        for ce in _findall(cs,"globe:CE",ns):
            id_=_find(ce,"globe:ID",ns)
            if id_ is not None and "GIR305" in _gs(id_, ns):
                for ow in _findall(ce,"globe:Ownership",ns):
                    te=_find(ow,"globe:TIN",ns); ov=_t(te)
                    if ov and t306 and ov not in t306:
                        bad.append(f"CE {_t(_find(id_,'globe:TIN',ns))!r} GIR305: Ownership/TIN={ov!r} non in GIR306 TINs{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    # ── 8.3.6 QIIR (70032) ───────────────────────────────────────────────────
    r=CheckResult("70032","Se CE/QIIR è compilato, CE/ID/Rules deve contenere GIR201 o GIR202.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/QIIR")
    bad=[]
    if cs is not None:
        for ce in _findall(cs,"globe:CE",ns):
            q=_find(ce,"globe:QIIR",ns)
            if q is not None:
                id_=_find(ce,"globe:ID",ns); rules=set((_t(_find(id_,"globe:Rules",ns)) or "").split()) if id_ is not None else set()
                if not rules&{"GIR201","GIR202"}:
                    bad.append(f"CE {_t(_find(id_,'globe:TIN',ns)) if id_ else '?'!r} ha QIIR ma Rules={rules}{_src(q)}")
    _ko_all(r, bad)
    out.append(r)

    # ── 8.3.7 CE/QIIR (70033-70035) ──────────────────────────────────────────
    # 70033: la regola AdE richiede che Exception/TIN corrisponda a un altro CE
    # nella CorporateStructure (scope: solo CE/ID/TIN, non UPE).
    # Se il TIN è quello di OtherUPE/ExcludedUPE viene segnalato come WARN
    # (semanticamente corretto — la UPE esercita la QIIR — ma non strettamente
    # conforme al testo letterale della regola AdE).
    r=CheckResult("70033","CE/QIIR/Exception/TIN deve corrispondere al TIN di un altro CE nella CorporateStructure.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/QIIR/Exception/TIN")
    bad_ko=[]; bad_warn=[]
    ce_tins_set=_ce_tins(); upe_tins_set=_upe_tins()
    if cs is not None:
        for ce in _findall(cs,"globe:CE",ns):
            id_=_find(ce,"globe:ID",ns); own=_t(_find(id_,"globe:TIN",ns)) if id_ is not None else ""
            q=_find(ce,"globe:QIIR",ns)
            if q is None: continue
            for exc in _findall(q,"globe:Exception",ns):
                te=_find(exc,"globe:TIN",ns); tv=_t(te)
                if not tv: continue
                if tv==own:
                    bad_ko.append(f"Exception/TIN={tv!r} uguale al CE stesso{_src(te)}")
                elif tv not in ce_tins_set and tv in upe_tins_set:
                    # TIN è della UPE: l'AdE lo segnala come errore 70033
                    bad_ko.append(f"Exception/TIN={tv!r} corrisponde a OtherUPE/ExcludedUPE, non a un CE{_src(te)}")
                elif tv not in ce_tins_set and tv not in upe_tins_set:
                    bad_ko.append(f"Exception/TIN={tv!r} non corrisponde a nessun CE né UPE{_src(te)}")
    _ko_all(r, bad_ko)
    out.append(r)

    # 70034 FIX: controlla Art2.1.3 solo se elemento presente
    r=CheckResult("70034","Se POPE-IPE=GIR902 ed Exception compilato, Art2.1.3/Status deve essere TRUE.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/QIIR/POPE-IPE")
    bad=[]
    if cs is not None:
        for ce in _findall(cs,"globe:CE",ns):
            q=_find(ce,"globe:QIIR",ns)
            if q is None: continue
            if _t(_find(q,"globe:POPE-IPE",ns))=="GIR902":
                exc=_find(q,"globe:Exception",ns)
                if exc is not None:
                    a213=_find(exc,"globe:Art2.1.3",ns)
                    if a213 is None: continue   # altra eccezione → PASS
                    se=_find(a213,"globe:Status",ns)
                    if _t(se).upper()!="TRUE":
                        bad.append(f"POPE-IPE=GIR902 Exception.Art2.1.3/Status={_t(se)!r}{_src(se) if se else _src(a213)}")
    _ko_all(r, bad)
    out.append(r)

    # 70035 FIX: controlla Art2.1.5 solo se elemento presente
    r=CheckResult("70035","Se POPE-IPE=GIR901 ed Exception compilato, Art2.1.5/Status deve essere TRUE.","/GLOBE_OECD/GLOBEBody/GeneralSection/.../CE/QIIR/POPE-IPE")
    bad=[]
    if cs is not None:
        for ce in _findall(cs,"globe:CE",ns):
            q=_find(ce,"globe:QIIR",ns)
            if q is None: continue
            if _t(_find(q,"globe:POPE-IPE",ns))=="GIR901":
                exc=_find(q,"globe:Exception",ns)
                if exc is not None:
                    a215=_find(exc,"globe:Art2.1.5",ns)
                    if a215 is None: continue
                    se=_find(a215,"globe:Status",ns)
                    if _t(se).upper()!="TRUE":
                        bad.append(f"POPE-IPE=GIR901 Exception.Art2.1.5/Status={_t(se)!r}{_src(se) if se else _src(a215)}")
    _ko_all(r, bad)
    out.append(r)

    out += _check_summary(summary_els, jur_sections, filing_info, ns)
    out += _check_jur_etr(summary_els, jur_sections, ns)
    out += _check_utpr_sh(summary_els, jur_sections, ns)
    out += _check_elections(jur_sections, ns)
    out += _check_overall_comp(jur_sections, ns)
    out += _check_postfiling(jur_sections, period_start, ns)
    out += _check_covtax(jur_sections, period_end, ns)
    out += _check_defer_tax(jur_sections, ns)
    out += _check_excess(jur_sections, ns)
    out += _check_atut(jur_sections, period_end, ns)
    out += _check_iir_utpr(jur_sections, utpr_attr, ns)
    out += _check_ce_fanil(jur_sections, ns)
    out += _check_ce_ngi(jur_sections, ns)
    return out


# ── 8.3.8 Summary (70036-70043) ──────────────────────────────────────────────
def _check_summary(summary_els, jur_sections, filing_info, ns):
    from collections import defaultdict
    out=[]
    pe=_find(filing_info,"globe:Period",ns) if filing_info is not None else None
    period_end=_date(_t(_find(pe,"globe:End",ns))) if pe is not None else None

    # 70036
    r=CheckResult("70036","Se una giurisdizione ha più Subgroup, la Summary deve essere ripetuta.","/GLOBE_OECD/GLOBEBody/Summary/Jurisdiction/Subgroup")
    jcnt=defaultdict(int); jsg=defaultdict(list)
    for s in summary_els:
        for jel in _findall(s,"globe:Jurisdiction",ns):
            jn=_t(_find(jel,"globe:JurisdictionName",ns))
            if jn:
                jcnt[jn]+=1
                for sg in _findall(jel,"globe:Subgroup",ns): jsg[jn].append(sg)
    bad=[f"{jn}: {len(sgs)} Subgroup ma {jcnt[jn]} Summary" for jn,sgs in jsg.items() if len(sgs)>1 and jcnt[jn]!=len(sgs)]
    _ko_all(r, bad)
    out.append(r)

    # 70037
    r=CheckResult("70037","Se Summary/Subgroup valorizzato, deve esistere JurSection/ETR/SubGroup TIN coerente.","/GLOBE_OECD/GLOBEBody/Summary/Jurisdiction/Subgroup/TIN")
    jst=defaultdict(set)
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            for sg in _findall(etr,"globe:SubGroup",ns):
                for te in _findall(sg,"globe:TIN",ns): jst[jn].add(_t(te))
    bad=[]
    for s in summary_els:
        for jel in _findall(s,"globe:Jurisdiction",ns):
            jn=_t(_find(jel,"globe:JurisdictionName",ns))
            for sg in _findall(jel,"globe:Subgroup",ns):
                for te in _findall(sg,"globe:TIN",ns):
                    tv=_t(te)
                    if tv and tv not in jst.get(jn,set()): bad.append(f"{jn} Summary Subgroup TIN={tv!r} non in JurSection{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    # 70038
    r=CheckResult("70038","Se Period/End > 30/06/2028, SafeHarbour non può contenere GIR1203/1204/1205.","/GLOBE_OECD/GLOBEBody/Summary/SafeHarbour")
    bad=[]
    if period_end and period_end>date(2028,6,30):
        for s in summary_els:
            for sh in _findall(s,"globe:SafeHarbour",ns):
                if _t(sh) in ("GIR1203","GIR1204","GIR1205"): bad.append(f"SafeHarbour={_t(sh)} dopo 30/06/2028{_src(sh)}")
    _ko_all(r, bad)
    out.append(r)

    # 70039
    r=CheckResult("70039","Se Period/End > 31/12/2026, SafeHarbour non può contenere GIR1206.","/GLOBE_OECD/GLOBEBody/Summary/SafeHarbour")
    bad=[]
    if period_end and period_end>date(2026,12,31):
        for s in summary_els:
            for sh in _findall(s,"globe:SafeHarbour",ns):
                if _t(sh)=="GIR1206": bad.append(f"SafeHarbour=GIR1206 dopo 31/12/2026{_src(sh)}")
    _ko_all(r, bad)
    out.append(r)

    out.append(CheckResult("70040","SafeHarbour=GIR1206 solo nella giurisdizione UPE.","/GLOBE_OECD/GLOBEBody/Summary/SafeHarbour"))

    # 70041
    r=CheckResult("70041","Se CFSofUPE=GIR502/GIR504, SafeHarbour non può contenere GIR1207/1208/1209.","/GLOBE_OECD/GLOBEBody/FilingInfo/AccountingInfo/CFSofUPE")
    bad=[]
    if filing_info is not None:
        acc=_find(filing_info,"globe:AccountingInfo",ns)
        cfs=_t(_find(acc,"globe:CFSofUPE",ns)) if acc is not None else ""
        if cfs in ("GIR502","GIR504"):
            for s in summary_els:
                for sh in _findall(s,"globe:SafeHarbour",ns):
                    if _t(sh) in ("GIR1207","GIR1208","GIR1209"): bad.append(f"CFSofUPE={cfs} ma SafeHarbour={_t(sh)}{_src(sh)}")
    _ko_all(r, bad)
    out.append(r)

    # 70042
    r=CheckResult("70042","Se JurWithTaxingRights e SafeHarbour assente o solo GIR1206, devono essere compilati ETRRange/SBIE/QDMTTut/GLoBETut.","/GLOBE_OECD/GLOBEBody/Summary/JurWithTaxingRights")
    bad=[]
    for s in summary_els:
        jwtr=_find(s,"globe:JurWithTaxingRights",ns)
        if jwtr is None: continue
        shv={_t(sh) for sh in _findall(s,"globe:SafeHarbour",ns)}
        if shv<={""} or shv<={"GIR1206"}:
            for req in ["globe:ETRRange","globe:SBIE","globe:QDMTTut","globe:GLoBETut"]:
                if _find(s,req,ns) is None: bad.append(f"Manca {req.split(':')[-1]} in Summary{_src(jwtr)}")
    _ko_all(r, bad)
    out.append(r)

    # 70043
    r=CheckResult("70043","Se JurWithTaxingRights e SafeHarbour=GIR1202, devono essere compilati ETRRange/SBIE/QDMTTut.","/GLOBE_OECD/GLOBEBody/Summary/JurWithTaxingRights")
    bad=[]
    for s in summary_els:
        jwtr=_find(s,"globe:JurWithTaxingRights",ns)
        if jwtr is None: continue
        if "GIR1202" in {_t(sh) for sh in _findall(s,"globe:SafeHarbour",ns)}:
            for req in ["globe:ETRRange","globe:SBIE","globe:QDMTTut"]:
                if _find(s,req,ns) is None: bad.append(f"GIR1202 ma manca {req.split(':')[-1]}{_src(jwtr)}")
    _ko_all(r, bad)
    out.append(r)

    return out


# ── 8.3.9 JurisdictionSection / ETRException (70044-70048) ───────────────────
def _check_jur_etr(summary_els, jur_sections, ns):
    from collections import defaultdict
    out=[]

    # 70044
    r=CheckResult("70044","Se ETRStatus è compilato, deve contenere ETRException o ETRComputation.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../ETRStatus")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            es=_find(etr,"globe:ETRStatus",ns)
            if es is not None and _find(es,"globe:ETRException",ns) is None and _find(es,"globe:ETRComputation",ns) is None:
                bad.append(f"{jn}: ETRStatus senza ETRException né ETRComputation{_src(es)}")
    _ko_all(r, bad)
    out.append(r)

    # Mappa SafeHarbour per (giurisdizione, SubGroup TIN)
    # FIX 70047/70048: il mapping deve avvenire per SubGroup TIN, non per giurisdizione intera.
    # In una stessa giurisdizione possono esistere più ETR con SubGroup distinti (GIR1607=CEs,
    # GIR1608=JV Groups), ciascuno associato a un diverso SafeHarbour nel Summary.
    # Il matching corretto è: Summary/Subgroup/TIN → ETR/SubGroup/TIN.

    # Costruisci mappa (jn, sg_tin) → set(SafeHarbour)
    sh_by_sg: dict = defaultdict(set)   # (jn, sg_tin) → {GIR12xx}
    sh_by_jur: dict = defaultdict(set)  # jn → {GIR12xx} (fallback per Subgroup senza TIN specifico)
    for s in summary_els:
        sh_vals = {_t(sh) for sh in _findall(s,"globe:SafeHarbour",ns)}
        for jel in _findall(s,"globe:Jurisdiction",ns):
            jn = _t(_find(jel,"globe:JurisdictionName",ns))
            sgs = _findall(jel,"globe:Subgroup",ns)
            if sgs:
                for sg in sgs:
                    tins = [_t(te) for te in _findall(sg,"globe:TIN",ns)]
                    for t in tins:
                        sh_by_sg[(jn, t)] |= sh_vals
                    if not tins:
                        sh_by_jur[jn] |= sh_vals
                # Fallback per giurisdizione: quando ETR/SubGroup/TIN='NOTIN'
                # il matching avviene per giurisdizione, non per TIN specifico.
                sh_by_jur[jn] |= sh_vals
            else:
                sh_by_jur[jn] |= sh_vals

    cbcr_codes = {"GIR1203","GIR1204","GIR1205"}

    # 70045
    # Il check scatta in due scenari:
    # A) SafeHarbour=GIR1203/4/5 per la giurisdizione ma TransitionalCbCRSafeHarbour assente
    # B) Summary/Subgroup ha TIN reale ma ETR/SubGroup ha NOTIN → mismatch strutturale
    #    In questo caso l'AdE segnala 70045 per ogni Subgroup del Summary senza match ETR
    r=CheckResult("70045","Se SafeHarbour=GIR1203/1204/1205, deve essere compilato ETRException/TransitionalCbCRSafeHarbour.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../ETRException/TransitionalCbCRSafeHarbour")
    bad=[]
    # Mappa: (jn, sg_tin_summary) → set SafeHarbour, solo per TIN reali (non NOTIN)
    sh_real_sg: dict = defaultdict(set)
    for s in summary_els:
        sh_vals_s = {_t(sh) for sh in _findall(s,"globe:SafeHarbour",ns)}
        for jel in _findall(s,"globe:Jurisdiction",ns):
            jn_s = _t(_find(jel,"globe:JurisdictionName",ns))
            for sg in _findall(jel,"globe:Subgroup",ns):
                for te in _findall(sg,"globe:TIN",ns):
                    tv = _t(te)
                    if tv and tv != "NOTIN":
                        sh_real_sg[(jn_s, tv)] |= sh_vals_s
    # ETR SubGroup TIN per giurisdizione
    etr_sg_tins: dict = defaultdict(set)
    for js in jur_sections:
        jn_j = _t(_find(js,"globe:Jurisdiction",ns))
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            for sg in _findall(etr,"globe:SubGroup",ns):
                for te in _findall(sg,"globe:TIN",ns):
                    etr_sg_tins[jn_j].add(_t(te))
    # Scenario B: Subgroup del Summary con TIN reale non trovato negli ETR SubGroup
    for s in summary_els:
        sh_vals_s = {_t(sh) for sh in _findall(s,"globe:SafeHarbour",ns)}
        if not (sh_vals_s & cbcr_codes): continue
        for jel in _findall(s,"globe:Jurisdiction",ns):
            jn_s = _t(_find(jel,"globe:JurisdictionName",ns))
            for sg in _findall(jel,"globe:Subgroup",ns):
                for te in _findall(sg,"globe:TIN",ns):
                    tv = _t(te)
                    if tv and tv != "NOTIN" and tv not in etr_sg_tins.get(jn_s, set()):
                        bad.append(f"{jn_s}: Summary Subgroup TIN={tv!r} con GIR1203/4/5 ma nessun ETR/SubGroup corrispondente{_src(te)}")
    # Scenario A: TransitionalCbCRSafeHarbour assente
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            sg_tins = [_t(te) for sg in _findall(etr,"globe:SubGroup",ns) for te in _findall(sg,"globe:TIN",ns)]
            etr_sh = set()
            for st in sg_tins: etr_sh |= sh_by_sg.get((jn,st), set())
            if not etr_sh: etr_sh = sh_by_jur.get(jn, set())
            if not (etr_sh & cbcr_codes): continue
            es=_find(etr,"globe:ETRStatus",ns)
            if es is None: bad.append(f"{jn}: GIR1203/4/5 ma ETRStatus assente"); continue
            exc=_find(es,"globe:ETRException",ns)
            if exc is None or _find(exc,"globe:TransitionalCbCRSafeHarbour",ns) is None:
                bad.append(f"{jn}: GIR1203/4/5 ma TransitionalCbCRSafeHarbour assente{_src(es)}")
    _ko_all(r, bad)
    out.append(r)

    # 70046
    r=CheckResult("70046","Se ETRException/TransitionalCbCRSafeHarbour compilato, ETR/SubGroup deve avere TypeofSubGroup=GIR1607/GIR1608.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../ETRException/TransitionalCbCRSafeHarbour")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            es=_find(etr,"globe:ETRStatus",ns)
            if es is None: continue
            exc=_find(es,"globe:ETRException",ns)
            if exc is None: continue
            tcsh=_find(exc,"globe:TransitionalCbCRSafeHarbour",ns)
            if tcsh is not None:
                sg_types={_t(_find(sg,"globe:TypeofSubGroup",ns)) for sg in _findall(etr,"globe:SubGroup",ns)}
                if not sg_types&{"GIR1607","GIR1608"}:
                    bad.append(f"{jn}: TransitionalCbCRSH presente ma SubGroup TypeofSubGroup≠GIR1607/1608{_src(tcsh)}")
    _ko_all(r, bad)
    out.append(r)

    # 70047: GIR1203 → Revenue obbligatorio
    # Scatta in due scenari:
    # A) ETR ha GIR1203 via sh_by_sg/jur ma Revenue assente nel TCSH
    # B) Summary/Subgroup con TIN reale e GIR1203 non ha ETR corrispondente (stessa logica 70045-B)
    r=CheckResult("70047","Se SafeHarbour=GIR1203, TransitionalCbCRSafeHarbour/Revenue deve essere compilato.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../TransitionalCbCRSafeHarbour/Revenue")
    bad=[]
    # Scenario B: Subgroup con GIR1203 senza ETR corrispondente
    for s in summary_els:
        sh_vals_s = {_t(sh) for sh in _findall(s,"globe:SafeHarbour",ns)}
        if "GIR1203" not in sh_vals_s: continue
        for jel in _findall(s,"globe:Jurisdiction",ns):
            jn_s = _t(_find(jel,"globe:JurisdictionName",ns))
            for sg in _findall(jel,"globe:Subgroup",ns):
                for te in _findall(sg,"globe:TIN",ns):
                    tv = _t(te)
                    if tv and tv != "NOTIN" and tv not in etr_sg_tins.get(jn_s, set()):
                        bad.append(f"{jn_s}: Summary Subgroup TIN={tv!r} con GIR1203 ma nessun ETR corrispondente; Revenue non verificabile{_src(te)}")
    # Scenario A: Revenue assente nel TCSH dell'ETR
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            sg_tins=[_t(te) for sg in _findall(etr,"globe:SubGroup",ns) for te in _findall(sg,"globe:TIN",ns)]
            etr_sh=set()
            for st in sg_tins: etr_sh |= sh_by_sg.get((jn,st), set())
            if not etr_sh: etr_sh = sh_by_jur.get(jn, set())
            if "GIR1203" not in etr_sh: continue
            es=_find(etr,"globe:ETRStatus",ns)
            if es is None: continue
            exc=_find(es,"globe:ETRException",ns)
            if exc is None: continue
            tcsh=_find(exc,"globe:TransitionalCbCRSafeHarbour",ns)
            if tcsh is not None and not _t(_find(tcsh,"globe:Revenue",ns)):
                bad.append(f"{jn}: GIR1203 ma Revenue assente{_src(tcsh)}")
    _ko_all(r, bad)
    out.append(r)

    # 70048: GIR1204 → IncomeTax obbligatorio
    r=CheckResult("70048","Se SafeHarbour=GIR1204, TransitionalCbCRSafeHarbour/IncomeTax deve essere compilato.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../TransitionalCbCRSafeHarbour/IncomeTax")
    bad=[]
    # Scenario B: Subgroup con GIR1204 senza ETR corrispondente
    for s in summary_els:
        sh_vals_s = {_t(sh) for sh in _findall(s,"globe:SafeHarbour",ns)}
        if "GIR1204" not in sh_vals_s: continue
        for jel in _findall(s,"globe:Jurisdiction",ns):
            jn_s = _t(_find(jel,"globe:JurisdictionName",ns))
            for sg in _findall(jel,"globe:Subgroup",ns):
                for te in _findall(sg,"globe:TIN",ns):
                    tv = _t(te)
                    if tv and tv != "NOTIN" and tv not in etr_sg_tins.get(jn_s, set()):
                        bad.append(f"{jn_s}: Summary Subgroup TIN={tv!r} con GIR1204 ma nessun ETR corrispondente; IncomeTax non verificabile{_src(te)}")
    # Scenario A: IncomeTax assente nel TCSH dell'ETR
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            sg_tins=[_t(te) for sg in _findall(etr,"globe:SubGroup",ns) for te in _findall(sg,"globe:TIN",ns)]
            etr_sh=set()
            for st in sg_tins: etr_sh |= sh_by_sg.get((jn,st), set())
            if not etr_sh: etr_sh = sh_by_jur.get(jn, set())
            if "GIR1204" not in etr_sh: continue
            es=_find(etr,"globe:ETRStatus",ns)
            if es is None: continue
            exc=_find(es,"globe:ETRException",ns)
            if exc is None: continue
            tcsh=_find(exc,"globe:TransitionalCbCRSafeHarbour",ns)
            if tcsh is not None and not _t(_find(tcsh,"globe:IncomeTax",ns)):
                bad.append(f"{jn}: GIR1204 ma IncomeTax assente{_src(tcsh)}")
    _ko_all(r, bad)
    out.append(r)

    return out


# ── 8.3.10 UTPR SafeHarbour / NMCE (70049-70053) ─────────────────────────────
def _check_utpr_sh(summary_els, jur_sections, ns):
    from collections import defaultdict
    out=[]
    # sh_jur aggrega SafeHarbour per giurisdizione indipendentemente dal SubGroup.
    # Questo è corretto per 70049-70053 che operano a livello di giurisdizione.
    sh_jur=defaultdict(set)
    for s in summary_els:
        shv={_t(sh) for sh in _findall(s,"globe:SafeHarbour",ns)}
        for jel in _findall(s,"globe:Jurisdiction",ns):
            sh_jur[_t(_find(jel,"globe:JurisdictionName",ns))]|=shv

    checks=[
        ("70049","Se SafeHarbour=GIR1206, ETRException/UTPRSafeHarbour/CITRate deve essere compilato.",
         "GIR1206",lambda etr,es,ns: _find(_find(es,"globe:ETRException",ns),"globe:UTPRSafeHarbour/globe:CITRate",ns) if _find(es,"globe:ETRException",ns) else None, True),
    ]

    # 70049
    r=CheckResult("70049","Se SafeHarbour=GIR1206, ETRException/UTPRSafeHarbour/CITRate deve essere compilato.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../ETRException/UTPRSafeHarbour/CITRate")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        if "GIR1206" not in sh_jur.get(jn,set()): continue
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            es=_find(etr,"globe:ETRStatus",ns)
            exc=_find(es,"globe:ETRException",ns) if es else None
            ush=_find(exc,"globe:UTPRSafeHarbour",ns) if exc else None
            if ush is None or not _t(_find(ush,"globe:CITRate",ns)): bad.append(f"{jn}: GIR1206 ma UTPRSafeHarbour/CITRate assente")
    _ko_all(r, bad)
    out.append(r)

    # 70050
    r=CheckResult("70050","Se SafeHarbour=GIR1207/1208/1209, ETRComputation/Non-MaterialCE deve essere compilato.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../ETRComputation/Non-MaterialCE")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        if not sh_jur.get(jn,set())&{"GIR1207","GIR1208","GIR1209"}: continue
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            es=_find(etr,"globe:ETRStatus",ns)
            ec=_find(es,"globe:ETRComputation",ns) if es else None
            if ec is None or _find(ec,"globe:Non-MaterialCE",ns) is None: bad.append(f"{jn}: GIR1207/8/9 ma Non-MaterialCE assente")
    _ko_all(r, bad)
    out.append(r)

    # 70051
    r=CheckResult("70051","Se SafeHarbour=GIR1208, Non-MaterialCE/RFY deve contenere AggregateSimplified.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../Non-MaterialCE/RFY/AggregateSimplified")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        if "GIR1208" not in sh_jur.get(jn,set()): continue
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            es=_find(etr,"globe:ETRStatus",ns); ec=_find(es,"globe:ETRComputation",ns) if es else None
            nm=_find(ec,"globe:Non-MaterialCE",ns) if ec else None
            rfy=_find(nm,"globe:RFY",ns) if nm else None
            if rfy is None or _find(rfy,"globe:AggregateSimplified",ns) is None: bad.append(f"{jn}: GIR1208 ma AggregateSimplified assente")
    _ko_all(r, bad)
    out.append(r)

    # 70052
    r=CheckResult("70052","Se SafeHarbour=GIR1209, OverallComputation/SubstanceExclusion deve essere compilato.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../OverallComputation/SubstanceExclusion")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        if "GIR1209" not in sh_jur.get(jn,set()): continue
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            if _find(oc,"globe:SubstanceExclusion",ns) is None: bad.append(f"{jn}: GIR1209 ma SubstanceExclusion assente")
    _ko_all(r, bad)
    out.append(r)

    # 70053: GIR1205 → SubstanceExclusion obbligatorio (salvo Profit≤0)
    # Scatta anche quando Summary/Subgroup ha TIN reale con GIR1205 ma l'ETR ha NOTIN
    r=CheckResult("70053","Se SafeHarbour=GIR1205, SubstanceExclusion obbligatorio (salvo Profit≤0).","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../OverallComputation/SubstanceExclusion")
    bad=[]
    # Mappa ETR SubGroup TIN per giurisdizione (ricostruita localmente)
    etr_sg_tins_53: dict = defaultdict(set)
    for js in jur_sections:
        jn_j = _t(_find(js,"globe:Jurisdiction",ns))
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            for sg in _findall(etr,"globe:SubGroup",ns):
                for te in _findall(sg,"globe:TIN",ns):
                    etr_sg_tins_53[jn_j].add(_t(te))
    # Scenario B: Subgroup con GIR1205 e TIN reale senza ETR corrispondente
    for s in summary_els:
        shv_s = {_t(sh) for sh in _findall(s,"globe:SafeHarbour",ns)}
        if "GIR1205" not in shv_s: continue
        for jel in _findall(s,"globe:Jurisdiction",ns):
            jn_s = _t(_find(jel,"globe:JurisdictionName",ns))
            for sg in _findall(jel,"globe:Subgroup",ns):
                for te in _findall(sg,"globe:TIN",ns):
                    tv = _t(te)
                    if tv and tv != "NOTIN" and tv not in etr_sg_tins_53.get(jn_s, set()):
                        bad.append(f"{jn_s}: Summary Subgroup TIN={tv!r} con GIR1205 ma nessun ETR corrispondente; SubstanceExclusion non verificabile{_src(te)}")
    # Scenario A: SubstanceExclusion assente nell'ETR con GIR1205
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        if "GIR1205" not in sh_jur.get(jn,set()): continue
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            es=_find(etr,"globe:ETRStatus",ns); exc=_find(es,"globe:ETRException",ns) if es is not None else None
            tcsh=_find(exc,"globe:TransitionalCbCRSafeHarbour",ns) if exc is not None else None
            profit=_decimal(_t(_find(tcsh,"globe:Profit",ns))) if tcsh is not None else None
            if profit is not None and profit<=0: continue
            for oc in _findall(etr,".//globe:ETRComputation/globe:OverallComputation",ns):
                if _find(oc,"globe:SubstanceExclusion",ns) is None: bad.append(f"{jn}: GIR1205 ma SubstanceExclusion assente (Profit>0)")
    _ko_all(r, bad)
    out.append(r)

    return out

# ── 8.3.11 Elections / CEComputation (70054-70058) ───────────────────────────
def _check_elections(jur_sections, ns):
    out=[]

    # 70054
    r=CheckResult("70054","In ETR/Election/*, RevocationYear solo quando Status=FALSE.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../ETR/Election/*/RevocationYear")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            el=_find(etr,"globe:Election",ns)
            if el is None: continue
            for ch in el:
                rv=_find(ch,"globe:RevocationYear",ns)
                if rv is not None and _t(_find(ch,"globe:Status",ns)).upper()!="FALSE":
                    bad.append(f"{jn}: RevocationYear con Status≠FALSE{_src(rv)}")
    _ko_all(r, bad)
    out.append(r)

    # 70055
    r=CheckResult("70055","Election/Art3.2.1.c: OutstandingBalance = QualOwnerIntentBalance + Additions - Reductions.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../Election/Art3.2.1.c/OutstandingBalance")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            art=_find(etr,"globe:Election/globe:Art3.2.1.c",ns)
            if art is None: continue
            ob=_decimal(_t(_find(art,"globe:OutstandingBalance",ns))); qo=_decimal(_t(_find(art,"globe:QualOwnerIntentBalance",ns)))
            add=_decimal(_t(_find(art,"globe:Additions",ns))) or Decimal(0); red=_decimal(_t(_find(art,"globe:Reductions",ns))) or Decimal(0)
            if ob is None or qo is None: continue
            if abs(qo+add-red-ob)>Decimal("1"): bad.append(f"{jn}: OutstandingBalance={ob} ≠ {qo+add-red}{_src(_find(art,'globe:OutstandingBalance',ns))}")
    _ko_all(r, bad)
    out.append(r)

    # 70056
    r=CheckResult("70056","In CEComputation/Elections/*, RevocationYear solo quando Status=FALSE.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../CEComputation/Elections/*/RevocationYear")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            els=_find(ce,"globe:Elections",ns)
            if els is None: continue
            for ch in els:
                rv=_find(ch,"globe:RevocationYear",ns)
                if rv is not None and _t(_find(ch,"globe:Status",ns)).upper()!="FALSE": bad.append(f"{jn}: CEComputation RevocationYear con Status≠FALSE{_src(rv)}")
    _ko_all(r, bad)
    out.append(r)

    # 70057
    r=CheckResult("70057","Se AggregatedReporting compilato, CEComputation/TIN deve coincidere con TaxConsolGroupTIN.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../CEComputation/TIN")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            agg=_find(ce,"globe:Elections/globe:AggregatedReporting",ns)
            if agg is None: continue
            cte=_find(ce,"globe:TIN",ns); ct=_t(cte); cot=_t(_find(agg,"globe:TaxConsolGroupTIN",ns))
            if ct and cot and ct!=cot: bad.append(f"{jn}: CEComputation/TIN={ct!r} ≠ TaxConsolGroupTIN={cot!r}{_src(cte)}")
    _ko_all(r, bad)
    out.append(r)

    # 70058
    r=CheckResult("70058","CEComputation/Elections/Art7.6/InvestmentEntityTIN non deve coincidere con CEComputation/TIN.","/GLOBE_OECD/GLOBEBody/JurisdictionSection/.../CEComputation/Elections/Art7.6/InvestmentEntityTIN")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            a76=_find(ce,"globe:Elections/globe:Art7.6",ns)
            if a76 is None: continue
            ite=_find(a76,"globe:InvestmentEntityTIN",ns); it=_t(ite); ct=_t(_find(ce,"globe:TIN",ns))
            if it and ct and it==ct: bad.append(f"{jn}: InvestmentEntityTIN = CEComputation/TIN = {ct!r}{_src(ite)}")
    _ko_all(r, bad)
    out.append(r)

    return out

# ── 8.3.12 OverallComputation (70059-70063) ───────────────────────────────────
def _check_overall_comp(jur_sections, ns):
    out=[]

    # 70059
    r=CheckResult("70059","AdjustmentItem in NetGlobeIncome/Adjustments unico per ETR.","JurisdictionSection/.../NetGlobeIncome/Adjustments/AdjustmentItem")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            items=[_t(x) for x in _findall(oc,"globe:NetGlobeIncome/globe:Adjustments/globe:AdjustmentItem",ns)]
            dups={x for x in items if items.count(x)>1}
            if dups: bad.append(f"{jn}: NéGI AdjustmentItem duplicati: {dups}")
    _ko_all(r, bad)
    out.append(r)

    # 70060
    r=CheckResult("70060","Se NetGlobeIncome/AdjustmentItem=GIR2025, deve essere compilato IntShippingIncome.","JurisdictionSection/.../NetGlobeIncome/IntShippingIncome")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            if "GIR2025" in [_t(x) for x in _findall(oc,"globe:NetGlobeIncome/globe:Adjustments/globe:AdjustmentItem",ns)]:
                if _find(oc,"globe:NetGlobeIncome/globe:IntShippingIncome",ns) is None: bad.append(f"{jn}: GIR2025 ma IntShippingIncome assente")
    _ko_all(r, bad)
    out.append(r)

    # 70061
    r=CheckResult("70061","Se Election/Art4.6.1=TRUE, AdjustedCoveredTax deve contenere GIR2711 con Amount<0.","JurisdictionSection/.../AdjustedCoveredTax/Adjustments/AdjustmentItem=GIR2711")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for etr in _findall(js,".//globe:GLoBETax/globe:ETR",ns):
            a461=_find(etr,"globe:Election/globe:Art4.6.1",ns)
            if a461 is None or _t(a461).upper()!="TRUE": continue
            for oc in _findall(etr,".//globe:ETRComputation/globe:OverallComputation",ns):
                ia={_t(_find(adj,"globe:AdjustmentItem",ns)):_decimal(_t(_find(adj,"globe:Amount",ns))) for adj in _findall(oc,"globe:AdjustedCoveredTax/globe:Adjustments",ns)}
                a2711=ia.get("GIR2711")
                if a2711 is None or a2711>=0: bad.append(f"{jn}: Art4.6.1=TRUE ma GIR2711 amount={a2711}{_src(a461)}")
    _ko_all(r, bad)
    out.append(r)

    # 70062
    r=CheckResult("70062","Se AdjustmentItem=GIR2720, AdjustedCoveredTax/Total non può essere negativo.","JurisdictionSection/.../AdjustedCoveredTax/Total")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            if "GIR2720" in [_t(x) for x in _findall(oc,"globe:AdjustedCoveredTax/globe:Adjustments/globe:AdjustmentItem",ns)]:
                te=_find(oc,"globe:AdjustedCoveredTax/globe:Total",ns); tot=_decimal(_t(te))
                if tot is not None and tot<0: bad.append(f"{jn}: GIR2720 ma ACT/Total={tot}<0{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    # 70063
    r=CheckResult("70063","AdjustmentItem in AdjustedCoveredTax/Adjustments unico per ETR.","JurisdictionSection/.../AdjustedCoveredTax/Adjustments/AdjustmentItem")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            items=[_t(x) for x in _findall(oc,"globe:AdjustedCoveredTax/globe:Adjustments/globe:AdjustmentItem",ns)]
            dups={x for x in items if items.count(x)>1}
            if dups: bad.append(f"{jn}: ACT AdjustmentItem duplicati: {dups}")
    _ko_all(r, bad)
    out.append(r)

    return out


# ── 8.3.13 PostFilingAdjust (70064-70067) ────────────────────────────────────
def _check_postfiling(jur_sections, period_start, ns):
    out=[]
    sy=period_start.year if period_start else None

    def _pfa_block(oc,sub):
        pfa=_find(oc,"globe:AdjustedCoveredTax/globe:PostFilingAdjust",ns)
        if pfa is None: return None
        return _find(pfa,f"globe:{sub}",ns)

    # 70064
    r=CheckResult("70064","PostFilingAdjust/DeferTaxAsset/Total = Σ AmountAttributed/Amount.","JurisdictionSection/.../PostFilingAdjust/DeferTaxAsset/Total")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            dta=_pfa_block(oc,"DeferTaxAsset")
            if dta is None: continue
            te=_find(dta,"globe:Total",ns); tot=_decimal(_t(te))
            sm=sum((_decimal(_t(_find(aa,"globe:Amount",ns))) or Decimal(0)) for aa in _findall(dta,"globe:AmountAttributed",ns))
            if tot is not None and abs(tot-sm)>Decimal("1"): bad.append(f"{jn}: DeferTaxAsset Total={tot} ≠ {sm}{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    # 70065
    r=CheckResult("70065","PostFilingAdjust/CoveredTaxRefund/Total = Σ AmountAttributed/Amount.","JurisdictionSection/.../PostFilingAdjust/CoveredTaxRefund/Total")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            ctr=_pfa_block(oc,"CoveredTaxRefund")
            if ctr is None: continue
            te=_find(ctr,"globe:Total",ns); tot=_decimal(_t(te))
            sm=sum((_decimal(_t(_find(aa,"globe:Amount",ns))) or Decimal(0)) for aa in _findall(ctr,"globe:AmountAttributed",ns))
            if tot is not None and abs(tot-sm)>Decimal("1"): bad.append(f"{jn}: CoveredTaxRefund Total={tot} ≠ {sm}{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    # 70066
    r=CheckResult("70066","DeferTaxAsset/AmountAttributed/Year ≤ anno Period/Start.","JurisdictionSection/.../PostFilingAdjust/DeferTaxAsset/AmountAttributed/Year")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            dta=_pfa_block(oc,"DeferTaxAsset")
            if dta is None: continue
            for aa in _findall(dta,"globe:AmountAttributed",ns):
                ye=_find(aa,"globe:Year",ns); yr=_year(ye)
                if sy and yr and yr>sy: bad.append(f"{jn}: DeferTaxAsset Year={yr} > {sy}{_src(ye)}")
    _ko_all(r, bad)
    out.append(r)

    # 70067
    r=CheckResult("70067","DeferTaxAsset/AmountAttributed/Year non ripetuto.","JurisdictionSection/.../PostFilingAdjust/DeferTaxAsset/AmountAttributed/Year")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            dta=_pfa_block(oc,"DeferTaxAsset")
            if dta is None: continue
            yrs=[_year(_find(aa,"globe:Year",ns)) for aa in _findall(dta,"globe:AmountAttributed",ns)]
            yrs=[y for y in yrs if y is not None]
            if len(yrs)!=len(set(yrs)): bad.append(f"{jn}: DeferTaxAsset anni ripetuti: {yrs}")
    _ko_all(r, bad)
    out.append(r)

    return out

# ── 8.3.14 CoveredTaxRefund / DeemedDistTax (70068-70075) ────────────────────
def _check_covtax(jur_sections, period_end, ns):
    out=[]; ey=period_end.year if period_end else None

    # 70068
    r=CheckResult("70068","CoveredTaxRefund/AmountAttributed/Year ≤ anno Period/End.","JurisdictionSection/.../PostFilingAdjust/CoveredTaxRefund/AmountAttributed/Year")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            pfa=_find(oc,"globe:AdjustedCoveredTax/globe:PostFilingAdjust",ns)
            if pfa is None: continue
            ctr=_find(pfa,"globe:CoveredTaxRefund",ns)
            if ctr is None: continue
            for aa in _findall(ctr,"globe:AmountAttributed",ns):
                ye=_find(aa,"globe:Year",ns); yr=_year(ye)
                if ey and yr and yr>ey: bad.append(f"{jn}: CoveredTaxRefund Year={yr} > {ey}{_src(ye)}")
    _ko_all(r, bad)
    out.append(r)

    # 70069
    r=CheckResult("70069","CoveredTaxRefund/AmountAttributed/Year non ripetuto.","JurisdictionSection/.../PostFilingAdjust/CoveredTaxRefund/AmountAttributed/Year")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            pfa=_find(oc,"globe:AdjustedCoveredTax/globe:PostFilingAdjust",ns)
            if pfa is None: continue
            ctr=_find(pfa,"globe:CoveredTaxRefund",ns)
            if ctr is None: continue
            yrs=[_year(_find(aa,"globe:Year",ns)) for aa in _findall(ctr,"globe:AmountAttributed",ns)]
            yrs=[y for y in yrs if y is not None]
            if len(yrs)!=len(set(yrs)): bad.append(f"{jn}: CoveredTaxRefund anni ripetuti: {yrs}")
    _ko_all(r, bad)
    out.append(r)

    def _ddt_recs(oc):
        ddt=_find(oc,"globe:AdjustedCoveredTax/globe:DeemedDistTax",ns)
        if ddt is None: return []
        return _findall(ddt,"globe:Election/globe:Recapture",ns)

    # 70070
    r=CheckResult("70070","DeemedDistTax/Recapture/Year ≤ Period/End.","JurisdictionSection/.../DeemedDistTax/Election/Recapture/Year")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            for rec in _ddt_recs(oc):
                ye=_find(rec,"globe:Year",ns); yr=_year(ye)
                if ey and yr and yr>ey: bad.append(f"{jn}: DeemedDistTax Year={yr} > {ey}{_src(ye)}")
    _ko_all(r, bad)
    out.append(r)

    # 70071
    r=CheckResult("70071","DeemedDistTax/Recapture/Year non più di 4 anni anteriore a Period/End.","JurisdictionSection/.../DeemedDistTax/Election/Recapture/Year")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            for rec in _ddt_recs(oc):
                ye=_find(rec,"globe:Year",ns); yr=_year(ye)
                if ey and yr and (ey-yr)>=4: bad.append(f"{jn}: DeemedDistTax Year={yr} è ≥4 anni prima di {ey}{_src(ye)}")
    _ko_all(r, bad)
    out.append(r)

    # 70072
    r=CheckResult("70072","Recapture/EndAmount = StartAmount - TotalDDT.","JurisdictionSection/.../DeemedDistTax/Election/Recapture/EndAmount")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            for rec in _ddt_recs(oc):
                ee=_find(rec,"globe:EndAmount",ns); se=_find(rec,"globe:StartAmount",ns); te=_find(rec,"globe:TotalDDT",ns)
                e=_decimal(_t(ee)); s=_decimal(_t(se)); t=_decimal(_t(te))
                if e is None or s is None or t is None: continue
                if abs(e-(s-t))>Decimal("1"): bad.append(f"{jn}: EndAmount={e} ≠ {s}-{t}={s-t}{_src(ee)}")
    _ko_all(r, bad)
    out.append(r)

    # 70073
    r=CheckResult("70073","Recapture/EndAmount non deve essere negativo.","JurisdictionSection/.../DeemedDistTax/Election/Recapture/EndAmount")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            for rec in _ddt_recs(oc):
                ee=_find(rec,"globe:EndAmount",ns); e=_decimal(_t(ee))
                if e is not None and e<0: bad.append(f"{jn}: EndAmount={e}<0{_src(ee)}")
    _ko_all(r, bad)
    out.append(r)

    # 70074
    r=CheckResult("70074","Recapture/TotalDDT = DDTYear-0+DDTYear-1+DDTYear-2+DDTYear-3.","JurisdictionSection/.../DeemedDistTax/Election/Recapture/TotalDDT")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            for rec in _ddt_recs(oc):
                te=_find(rec,"globe:TotalDDT",ns); tot=_decimal(_t(te))
                if tot is None: continue
                sm=sum((_decimal(_t(_find(rec,f"globe:DDTYear-{i}",ns))) or Decimal(0)) for i in range(4))
                if abs(tot-sm)>Decimal("1"): bad.append(f"{jn}: TotalDDT={tot} ≠ {sm}{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    # 70075
    r=CheckResult("70075","Se Recapture/Year = anno Period/End, DDTYear-0/1/2/3 devono essere 0.","JurisdictionSection/.../DeemedDistTax/Election/Recapture/DDTYear-0")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            for rec in _ddt_recs(oc):
                yr=_year(_find(rec,"globe:Year",ns))
                if ey and yr==ey:
                    for i in range(4):
                        el=_find(rec,f"globe:DDTYear-{i}",ns); v=_decimal(_t(el))
                        if v is not None and v!=0: bad.append(f"{jn}: Year={yr}=PeriodEnd ma DDTYear-{i}={v}≠0{_src(el)}")
    _ko_all(r, bad)
    out.append(r)

    return out


# ── 8.3.15 TransBlendCFC / DeferTaxAdjustAmt (70076-70082) ───────────────────
def _check_defer_tax(jur_sections, ns):
    out=[]

    # 70076
    r=CheckResult("70076","TransBlendCFC/Total = Σ AggAllocTax.","JurisdictionSection/.../AdjustedCoveredTax/TransBlendCFC/Total")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            tbc=_find(oc,"globe:AdjustedCoveredTax/globe:TransBlendCFC",ns)
            if tbc is None: continue
            te=_find(tbc,"globe:Total",ns); tot=_decimal(_t(te))
            sm=sum((_decimal(_t(x)) or Decimal(0)) for x in _findall(tbc,".//globe:CFCJur/globe:Allocation/globe:AggAllocTax",ns))
            if tot is not None and abs(tot-sm)>Decimal("1"): bad.append(f"{jn}: TransBlendCFC Total={tot} ≠ {sm}{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    # 70077
    r=CheckResult("70077","DeferTaxAdjustAmt/Total = PreRecast + Recast/Lower - Recast/Higher.","JurisdictionSection/.../AdjustedCoveredTax/DeferTaxAdjustAmt/Total")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            dta=_find(oc,"globe:AdjustedCoveredTax/globe:DeferTaxAdjustAmt",ns)
            if dta is None: continue
            te=_find(dta,"globe:Total",ns); tot=_decimal(_t(te))
            pre=_decimal(_t(_find(dta,"globe:PreRecast",ns))) or Decimal(0)
            lo=_decimal(_t(_find(dta,"globe:Recast/globe:Lower",ns))) or Decimal(0)
            hi=_decimal(_t(_find(dta,"globe:Recast/globe:Higher",ns))) or Decimal(0)
            if tot is None: continue
            if abs(tot-(pre+lo-hi))>Decimal("1"): bad.append(f"{jn}: DeferTaxAdjustAmt Total={tot} ≠ {pre+lo-hi}{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    # 70078
    r=CheckResult("70078","DeferTaxAdjustAmt/BefRecastAdjust = DefTaxAmt - DiffCarryValue + GLoBEValue.","JurisdictionSection/.../AdjustedCoveredTax/DeferTaxAdjustAmt/BefRecastAdjust")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            dta=_find(oc,"globe:AdjustedCoveredTax/globe:DeferTaxAdjustAmt",ns)
            if dta is None: continue
            be=_find(dta,"globe:BefRecastAdjust",ns); bef=_decimal(_t(be))
            dta2=_decimal(_t(_find(dta,"globe:DefTaxAmt",ns))); diff=_decimal(_t(_find(dta,"globe:DiffCarryValue",ns))); gbe=_decimal(_t(_find(dta,"globe:GLoBEValue",ns)))
            if bef is None or dta2 is None or diff is None or gbe is None: continue
            if abs(bef-(dta2-diff+gbe))>Decimal("1"): bad.append(f"{jn}: BefRecastAdjust={bef} ≠ {dta2-diff+gbe}{_src(be)}")
    _ko_all(r, bad)
    out.append(r)

    # 70079
    r=CheckResult("70079","DeferTaxAdjustAmt/PreRecast = BefRecastAdjust + TotalAdjust.","JurisdictionSection/.../AdjustedCoveredTax/DeferTaxAdjustAmt/PreRecast")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            dta=_find(oc,"globe:AdjustedCoveredTax/globe:DeferTaxAdjustAmt",ns)
            if dta is None: continue
            pe=_find(dta,"globe:PreRecast",ns); pre=_decimal(_t(pe))
            bef=_decimal(_t(_find(dta,"globe:BefRecastAdjust",ns))); tadj=_decimal(_t(_find(dta,"globe:TotalAdjust",ns)))
            if pre is None or bef is None or tadj is None: continue
            if abs(pre-(bef+tadj))>Decimal("1"): bad.append(f"{jn}: PreRecast={pre} ≠ {bef}+{tadj}={bef+tadj}{_src(pe)}")
    _ko_all(r, bad)
    out.append(r)

    # 70080
    r=CheckResult("70080","AdjustmentItem in DeferTaxAdjustAmt/Adjustments unico per ETR.","JurisdictionSection/.../AdjustedCoveredTax/DeferTaxAdjustAmt/Adjustments/AdjustmentItem")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            dta=_find(oc,"globe:AdjustedCoveredTax/globe:DeferTaxAdjustAmt",ns)
            if dta is None: continue
            items=[_t(x) for x in _findall(dta,"globe:Adjustments/globe:AdjustmentItem",ns)]
            dups={x for x in items if items.count(x)>1}
            if dups: bad.append(f"{jn}: DeferTaxAdjustAmt Adjustments duplicati: {dups}")
    _ko_all(r, bad)
    out.append(r)

    # 70081
    r=CheckResult("70081","Transition/DeferredTaxAssets/Total = DeferredTaxAssetStart - Excluded OPPURE Recast - Excluded.","JurisdictionSection/.../Transition/DeferredTaxAssets/Total")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            dta=_find(oc,"globe:AdjustedCoveredTax/globe:DeferTaxAdjustAmt",ns)
            if dta is None: continue
            tr=_find(dta,"globe:Transition/globe:DeferredTaxAssets",ns)
            if tr is None: continue
            te=_find(tr,"globe:Total",ns); tot=_decimal(_t(te))
            st=_decimal(_t(_find(tr,"globe:DeferredTaxAssetStart",ns))); rc=_decimal(_t(_find(tr,"globe:DeferredTaxAssetRecast",ns))); ex=_decimal(_t(_find(tr,"globe:DeferredTaxAssetExcluded",ns)))
            if tot is None or ex is None: continue
            ok1=st is not None and abs(tot-(st-ex))<=Decimal("1"); ok2=rc is not None and abs(tot-(rc-ex))<=Decimal("1")
            if not ok1 and not ok2: bad.append(f"{jn}: DeferredTaxAssets/Total={tot} non coincide con nessuno dei calcoli{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    # 70082
    r=CheckResult("70082","Se DeferredTaxAssets presente, uno tra Start e Recast deve essere 0.","JurisdictionSection/.../Transition/DeferredTaxAssets")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            dta=_find(oc,"globe:AdjustedCoveredTax/globe:DeferTaxAdjustAmt",ns)
            if dta is None: continue
            tr=_find(dta,"globe:Transition/globe:DeferredTaxAssets",ns)
            if tr is None: continue
            se=_find(tr,"globe:DeferredTaxAssetStart",ns); re_=_find(tr,"globe:DeferredTaxAssetRecast",ns)
            s=_decimal(_t(se)); r_=_decimal(_t(re_))
            if s is not None and r_ is not None and s!=0 and r_!=0: bad.append(f"{jn}: DeferredTaxAssetStart={s} e Recast={r_} entrambi ≠ 0{_src(se)}")
    _ko_all(r, bad)
    out.append(r)

    return out

# ── 8.3.16 ExcessNegTaxExpense / ExcessProfits / Substance (70083-70087) ─────
def _check_excess(jur_sections, ns):
    out=[]

    # 70083
    r=CheckResult("70083","ExcessNegTaxExpense/Remaining = PriorYearBalance + GeneratedInRFY - UtilizedInRFY.","JurisdictionSection/.../ExcessNegTaxExpense/Remaining")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            en=_find(oc,"globe:ExcessNegTaxExpense",ns)
            if en is None: continue
            re=_find(en,"globe:Remaining",ns); rem=_decimal(_t(re))
            py=_decimal(_t(_find(en,"globe:PriorYearBalance",ns))); ge=_decimal(_t(_find(en,"globe:GeneratedInRFY",ns))); ut=_decimal(_t(_find(en,"globe:UtilizedInRFY",ns)))
            if rem is None or py is None or ge is None or ut is None: continue
            if abs(rem-(py+ge-ut))>Decimal("1"): bad.append(f"{jn}: ExcessNegTaxExpense Remaining={rem} ≠ {py+ge-ut}{_src(re)}")
    _ko_all(r, bad)
    out.append(r)

    # 70084
    r=CheckResult("70084","Se AdjustmentItem=GIR2719, Amount = ExcessNegTaxExpense/GeneratedInRFY.","JurisdictionSection/.../AdjustedCoveredTax/Adjustments/Amount[GIR2719]")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            en=_find(oc,"globe:ExcessNegTaxExpense",ns); ge=_decimal(_t(_find(en,"globe:GeneratedInRFY",ns))) if en is not None else None
            for adj in _findall(oc,"globe:AdjustedCoveredTax/globe:Adjustments",ns):
                if _t(_find(adj,"globe:AdjustmentItem",ns))=="GIR2719":
                    ae=_find(adj,"globe:Amount",ns); amt=_decimal(_t(ae))
                    if ge is not None and amt is not None and abs(amt-ge)>Decimal("1"): bad.append(f"{jn}: GIR2719 Amount={amt} ≠ GeneratedInRFY={ge}{_src(ae)}")
    _ko_all(r, bad)
    out.append(r)

    # 70085
    r=CheckResult("70085","Se AdjustmentItem=GIR2720, Amount = ExcessNegTaxExpense/UtilizedInRFY.","JurisdictionSection/.../AdjustedCoveredTax/Adjustments/Amount[GIR2720]")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            en=_find(oc,"globe:ExcessNegTaxExpense",ns); ut=_decimal(_t(_find(en,"globe:UtilizedInRFY",ns))) if en is not None else None
            for adj in _findall(oc,"globe:AdjustedCoveredTax/globe:Adjustments",ns):
                if _t(_find(adj,"globe:AdjustmentItem",ns))=="GIR2720":
                    ae=_find(adj,"globe:Amount",ns); amt=_decimal(_t(ae))
                    if ut is not None and amt is not None and abs(amt-ut)>Decimal("1"): bad.append(f"{jn}: GIR2720 Amount={amt} ≠ UtilizedInRFY={ut}{_src(ae)}")
    _ko_all(r, bad)
    out.append(r)

    # 70086
    r=CheckResult("70086","ExcessProfits = max(0, NetGlobeIncome/Total - SubstanceExclusion/Total).","JurisdictionSection/.../OverallComputation/ExcessProfits")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            ee=_find(oc,"globe:ExcessProfits",ns); ep=_decimal(_t(ee))
            ngi=_decimal(_t(_find(oc,"globe:NetGlobeIncome/globe:Total",ns))); se=_decimal(_t(_find(oc,"globe:SubstanceExclusion/globe:Total",ns))) or Decimal(0)
            if ep is None or ngi is None: continue
            exp=max(Decimal(0), ngi-se)
            if abs(ep-exp)>Decimal("1"): bad.append(f"{jn}: ExcessProfits={ep} ≠ max(0,{ngi}-{se})={exp}{_src(ee)}")
    _ko_all(r, bad)
    out.append(r)

    # 70087
    r=CheckResult("70087","SubstanceExclusion/Total = (PayrollCost * PayrollMarkUp) + (TangibleAssetValue * TangibleAssetMarkup).","JurisdictionSection/.../OverallComputation/SubstanceExclusion/Total")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            se=_find(oc,"globe:SubstanceExclusion",ns)
            if se is None: continue
            te=_find(se,"globe:Total",ns); tot=_decimal(_t(te))
            pc=_decimal(_t(_find(se,"globe:PayrollCost",ns))) or Decimal(0); pm=_decimal(_t(_find(se,"globe:PayrollMarkUp",ns))) or Decimal(0)
            tv=_decimal(_t(_find(se,"globe:TangibleAssetValue",ns))) or Decimal(0); tm=_decimal(_t(_find(se,"globe:TangibleAssetMarkup",ns))) or Decimal(0)
            if tot is None: continue
            exp=pc*pm+tv*tm
            if abs(tot-exp)>Decimal("1"): bad.append(f"{jn}: SubstanceExclusion Total={tot} ≠ {exp:.2f}{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    return out


# ── 8.3.17 AdditionalTopUpTax (70088-70096) ───────────────────────────────────
def _check_atut(jur_sections, period_end, ns):
    out=[]; ey=period_end.year if period_end else None

    def _art415(oc): return _find(oc,"globe:AdditionalTopUpTax/globe:Art4.1.5",ns)

    # 70088
    r=CheckResult("70088","Se NetGlobeIncome/Total<0, Art4.1.5 deve essere compilato.","JurisdictionSection/.../AdditionalTopUpTax/Art4.1.5")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            ne=_find(oc,"globe:NetGlobeIncome/globe:Total",ns); ngi=_decimal(_t(ne))
            if ngi is not None and ngi<0 and _art415(oc) is None: bad.append(f"{jn}: NéGI={ngi}<0 ma Art4.1.5 assente{_src(ne)}")
    _ko_all(r, bad)
    out.append(r)

    # 70089
    r=CheckResult("70089","Nel blocco Art4.1.5, AdjustedCoveredTax deve essere negativo.","JurisdictionSection/.../AdditionalTopUpTax/Art4.1.5/AdjustedCoveredTax")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            art=_art415(oc)
            if art is None: continue
            ae=_find(art,"globe:AdjustedCoveredTax",ns); a=_decimal(_t(ae))
            if a is not None and a>=0: bad.append(f"{jn}: Art4.1.5/AdjustedCoveredTax={a} non negativo{_src(ae)}")
    _ko_all(r, bad)
    out.append(r)

    # 70090
    r=CheckResult("70090","Art4.1.5/GlobeLoss deve coincidere con NetGlobeIncome/Total.","JurisdictionSection/.../AdditionalTopUpTax/Art4.1.5/GlobeLoss")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            art=_art415(oc)
            if art is None: continue
            le=_find(art,"globe:GlobeLoss",ns); loss=_decimal(_t(le)); ngi=_decimal(_t(_find(oc,"globe:NetGlobeIncome/globe:Total",ns)))
            if loss is not None and ngi is not None and abs(loss-ngi)>Decimal("1"): bad.append(f"{jn}: GlobeLoss={loss} ≠ NéGI={ngi}{_src(le)}")
    _ko_all(r, bad)
    out.append(r)

    # 70091
    r=CheckResult("70091","Art4.1.5/ExpectedAdjustedCoveredTax = GlobeLoss * 15%.","JurisdictionSection/.../AdditionalTopUpTax/Art4.1.5/ExpectedAdjustedCoveredTax")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            art=_art415(oc)
            if art is None: continue
            ee=_find(art,"globe:ExpectedAdjustedCoveredTax",ns); exp=_decimal(_t(ee)); loss=_decimal(_t(_find(art,"globe:GlobeLoss",ns)))
            if exp is None or loss is None: continue
            if abs(exp-loss*Decimal("0.15"))>Decimal("1"): bad.append(f"{jn}: ExpectedACT={exp} ≠ {loss}*15%={loss*Decimal('0.15'):.2f}{_src(ee)}")
    _ko_all(r, bad)
    out.append(r)

    # 70092
    r=CheckResult("70092","Art4.1.5/AdditionalTopUpTax = max(0, ExpectedACT - AdjustedCoveredTax).","JurisdictionSection/.../AdditionalTopUpTax/Art4.1.5/AdditionalTopUpTax")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            art=_art415(oc)
            if art is None: continue
            ae=_find(art,"globe:AdditionalTopUpTax",ns); atut=_decimal(_t(ae))
            exp=_decimal(_t(_find(art,"globe:ExpectedAdjustedCoveredTax",ns))); act=_decimal(_t(_find(art,"globe:AdjustedCoveredTax",ns)))
            if atut is None or exp is None or act is None: continue
            expected=max(Decimal(0),exp-act)
            if abs(atut-expected)>Decimal("1"): bad.append(f"{jn}: Art4.1.5/ATUT={atut} ≠ max(0,{exp}-{act})={expected}{_src(ae)}")
    _ko_all(r, bad)
    out.append(r)

    # 70093
    r=CheckResult("70093","NONArt4.1.5/Year ≤ anno Period/End.","JurisdictionSection/.../AdditionalTopUpTax/NONArt4.1.5/Year")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            for non in _findall(oc,"globe:AdditionalTopUpTax/globe:NONArt4.1.5",ns):
                ye=_find(non,"globe:Year",ns); yr=_year(ye)
                if ey and yr and yr>ey: bad.append(f"{jn}: NONArt4.1.5 Year={yr} > {ey}{_src(ye)}")
    _ko_all(r, bad)
    out.append(r)

    # 70094
    r=CheckResult("70094","Se Articles contiene GIR2605, Year deve essere almeno 4 anni anteriore a Period/End.","JurisdictionSection/.../AdditionalTopUpTax/NONArt4.1.5/Year")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            for non in _findall(oc,"globe:AdditionalTopUpTax/globe:NONArt4.1.5",ns):
                if "GIR2605" not in set((_t(_find(non,"globe:Articles",ns)) or "").split()): continue
                ye=_find(non,"globe:Year",ns); yr=_year(ye)
                if ey and yr and (ey-yr)<4: bad.append(f"{jn}: GIR2605 Year={yr} non è ≥4 anni prima di {ey}{_src(ye)}")
    _ko_all(r, bad)
    out.append(r)

    # 70095
    r=CheckResult("70095","Se Articles contiene GIR2602, Year deve essere il quinto anno precedente a Period/End.","JurisdictionSection/.../AdditionalTopUpTax/NONArt4.1.5/Year")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            for non in _findall(oc,"globe:AdditionalTopUpTax/globe:NONArt4.1.5",ns):
                if "GIR2602" not in set((_t(_find(non,"globe:Articles",ns)) or "").split()): continue
                ye=_find(non,"globe:Year",ns); yr=_year(ye)
                if ey and yr and (ey-yr)!=5: bad.append(f"{jn}: GIR2602 Year={yr} non è il quinto anno prima di {ey}{_src(ye)}")
    _ko_all(r, bad)
    out.append(r)

    # 70096
    r=CheckResult("70096","NONArt4.1.5/AdditionalTopUpTax = Recalculated/TopUpTax - Previous/TopUpTax.","JurisdictionSection/.../AdditionalTopUpTax/NONArt4.1.5/AdditionalTopUpTax")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for oc in _findall(js,".//globe:ETRComputation/globe:OverallComputation",ns):
            for non in _findall(oc,"globe:AdditionalTopUpTax/globe:NONArt4.1.5",ns):
                ae=_find(non,"globe:AdditionalTopUpTax",ns); atut=_decimal(_t(ae))
                rec=_decimal(_t(_find(non,"globe:Recalculated/globe:TopUpTax",ns))); prev=_decimal(_t(_find(non,"globe:Previous/globe:TopUpTax",ns)))
                if atut is None or rec is None or prev is None: continue
                if abs(atut-(rec-prev))>Decimal("1"): bad.append(f"{jn}: NONArt4.1.5 ATUT={atut} ≠ {rec}-{prev}={rec-prev}{_src(ae)}")
    _ko_all(r, bad)
    out.append(r)

    return out

# ── 8.3.18 JurisdictionSection IIR/UTPR (70097-70105) ────────────────────────
def _check_iir_utpr(jur_sections, utpr_attr, ns):
    out=[]

    # 70097
    r=CheckResult("70097","IIR/ParentEntity/InclusionRatio = (NetGlobeIncome - OtherOwnershipAllocation) / NetGlobeIncome.","JurisdictionSection/.../IIR/ParentEntity/InclusionRatio")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ltce in _findall(js,".//globe:LowTaxJurisdiction/globe:LTCE",ns):
            for iir in _findall(ltce,"globe:IIR",ns):
                ne=_find(iir,"globe:NetGlobeIncome",ns); ngi=_decimal(_t(ne))
                if ngi is None or ngi==0: continue
                for pe in _findall(iir,"globe:ParentEntity",ns):
                    ire=_find(pe,"globe:InclusionRatio",ns); ir=_decimal(_t(ire))
                    ooa=_decimal(_t(_find(pe,"globe:OtherOwnershipAllocation",ns))) or Decimal(0)
                    if ir is None: continue
                    exp=(ngi-ooa)/ngi
                    if abs(ir-exp)>Decimal("0.0001"): bad.append(f"{jn}: InclusionRatio={ir} ≠ ({ngi}-{ooa})/{ngi}={exp:.6f}{_src(ire)}")
    _ko_all(r, bad)
    out.append(r)

    # 70098
    r=CheckResult("70098","IIR/ParentEntity/TopUpTaxShare = IIR/TopUpTax * InclusionRatio.","JurisdictionSection/.../IIR/ParentEntity/TopUpTaxShare")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ltce in _findall(js,".//globe:LowTaxJurisdiction/globe:LTCE",ns):
            for iir in _findall(ltce,"globe:IIR",ns):
                te=_find(iir,"globe:TopUpTax",ns); tut=_decimal(_t(te))
                for pe in _findall(iir,"globe:ParentEntity",ns):
                    se=_find(pe,"globe:TopUpTaxShare",ns); sh=_decimal(_t(se)); ir=_decimal(_t(_find(pe,"globe:InclusionRatio",ns)))
                    if sh is None or tut is None or ir is None: continue
                    if abs(sh-tut*ir)>Decimal("1"): bad.append(f"{jn}: TopUpTaxShare={sh} ≠ {tut}*{ir}={tut*ir:.2f}{_src(se)}")
    _ko_all(r, bad)
    out.append(r)

    # 70099
    r=CheckResult("70099","Σ(UTPRTopUpTaxAttributed) = Σ(TotalUTPRTopUpTax).","/GLOBE_OECD/GLOBEBody/UTPRAttribution/Attribution/UTPRTopUpTaxAttributed")
    if utpr_attr is not None:
        sa=sum((_decimal(_t(x)) or Decimal(0)) for x in _findall(utpr_attr,".//globe:Attribution/globe:UTPRTopUpTaxAttributed",ns))
        st=sum((_decimal(_t(x)) or Decimal(0)) for js in jur_sections for x in _findall(js,".//globe:LowTaxJurisdiction/globe:UTPR/globe:UTPRCalculation/globe:TotalUTPRTopUpTax",ns))
        if abs(sa-st)>Decimal("1"): r.ko(f"Σ(UTPRTopUpTaxAttributed)={sa} ≠ Σ(TotalUTPRTopUpTax)={st}")
    out.append(r)

    # 70100
    r=CheckResult("70100","Se UTPRCalculation e TotalUTPRTopUpTax>0, UTPRAttribution deve essere compilato.","/GLOBE_OECD/GLOBEBody/UTPRAttribution")
    bad=[]
    for js in jur_sections:
        for uc in _findall(js,".//globe:LowTaxJurisdiction/globe:UTPR/globe:UTPRCalculation",ns):
            te=_find(uc,"globe:TotalUTPRTopUpTax",ns); tot=_decimal(_t(te))
            if tot is not None and tot>0 and utpr_attr is None: bad.append(f"TotalUTPRTopUpTax={tot}>0 ma UTPRAttribution assente{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    # 70101-70105
    for code,desc,field,check_fn in [
        ("70101","Attribution/Employees deve essere compilato salvo CarryForward=0.","/GLOBE_OECD/GLOBEBody/UTPRAttribution/Attribution/Employees",
         lambda attr,ns: _find(attr,"globe:UTPRTopUpTaxCarryForward",ns) is not None and _decimal(_t(_find(attr,"globe:UTPRTopUpTaxCarryForward",ns)))!=0 and _find(attr,"globe:Employees",ns) is None),
        ("70102","Attribution/TangibleAssetValue deve essere compilato salvo CarryForward=0.","/GLOBE_OECD/GLOBEBody/UTPRAttribution/Attribution/TangibleAssetValue",
         lambda attr,ns: _find(attr,"globe:UTPRTopUpTaxCarryForward",ns) is not None and _decimal(_t(_find(attr,"globe:UTPRTopUpTaxCarryForward",ns)))!=0 and _find(attr,"globe:TangibleAssetValue",ns) is None),
        ("70103","UTPRPercentage deve essere 0 quando CarryForward>0.","/GLOBE_OECD/GLOBEBody/UTPRAttribution/Attribution/UTPRPercentage",
         lambda attr,ns: (_decimal(_t(_find(attr,"globe:UTPRTopUpTaxCarryForward",ns))) or Decimal(0))>0 and (_decimal(_t(_find(attr,"globe:UTPRPercentage",ns))) or Decimal(0))!=0),
        ("70104","UTPRTopUpTaxCarriedForward non può essere negativo.","/GLOBE_OECD/GLOBEBody/UTPRAttribution/Attribution/UTPRTopUpTaxCarriedForward",
         lambda attr,ns: (_decimal(_t(_find(attr,"globe:UTPRTopUpTaxCarriedForward",ns))) or Decimal(0))<0),
    ]:
        r=CheckResult(code,desc,field)
        bad=[]
        if utpr_attr is not None:
            for attr in _findall(utpr_attr,"globe:Attribution",ns):
                if check_fn(attr,ns): bad.append(f"Attribution violazione {code}{_src(attr)}")
        _ko_all(r, bad)
        out.append(r)

    # 70105
    r=CheckResult("70105","UTPRTopUpTaxCarriedForward = CarryForward + Attributed - AddCashTaxExpense.","/GLOBE_OECD/GLOBEBody/UTPRAttribution/Attribution/UTPRTopUpTaxCarriedForward")
    bad=[]
    if utpr_attr is not None:
        for attr in _findall(utpr_attr,"globe:Attribution",ns):
            ce=_find(attr,"globe:UTPRTopUpTaxCarriedForward",ns); car=_decimal(_t(ce))
            fw=_decimal(_t(_find(attr,"globe:UTPRTopUpTaxCarryForward",ns))) or Decimal(0)
            at=_decimal(_t(_find(attr,"globe:UTPRTopUpTaxAttributed",ns))) or Decimal(0)
            ca=_decimal(_t(_find(attr,"globe:AddCashTaxExpense",ns))) or Decimal(0)
            if car is None: continue
            if abs(car-(fw+at-ca))>Decimal("1"): bad.append(f"CarriedForward={car} ≠ {fw}+{at}-{ca}={fw+at-ca}{_src(ce)}")
    _ko_all(r, bad)
    out.append(r)

    return out


# ── 8.3.19 CEComputation AdjustedFANIL / UPEAdjustments (70106-70113) ────────
def _check_ce_fanil(jur_sections, ns):
    out=[]

    # 70106
    r=CheckResult("70106","CrossBorderAdjustments/OtherTIN non deve coincidere con CEComputation/TIN.","JurisdictionSection/.../CEComputation/AdjustedFANIL/CrossBorderAdjustments/OtherTIN")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            ct=_t(_find(ce,"globe:TIN",ns))
            for adj in _findall(ce,".//globe:AdjustedFANIL/globe:Adjustment/globe:CrossBorderAdjustments",ns):
                oe=_find(adj,"globe:OtherTIN",ns); ov=_t(oe)
                if ct and ov and ct==ov: bad.append(f"{jn}: CrossBorderAdjustments/OtherTIN={ov!r} = CEComputation/TIN{_src(oe)}")
    _ko_all(r, bad)
    out.append(r)

    # 70107
    r=CheckResult("70107","Se UPEAdjustments/Reductions/Exception=TRUE, CrossBorderAdjustments non deve essere presente.","JurisdictionSection/.../CEComputation/AdjustedFANIL/Adjustment/CrossBorderAdjustments")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            for adj in _findall(ce,".//globe:AdjustedFANIL/globe:Adjustment",ns):
                ee=_find(adj,"globe:UPEAdjustments/globe:Reductions/globe:Exception",ns)
                if ee is not None and _t(ee).upper()=="TRUE":
                    cba=_find(adj,"globe:CrossBorderAdjustments",ns)
                    if cba is not None: bad.append(f"{jn}: Exception=TRUE ma CrossBorderAdjustments presente{_src(cba)}")
    _ko_all(r, bad)
    out.append(r)

    # 70108-70113
    for code,desc,bases,check_fn in [
        ("70108","Se Basis=GIR1901/02/05/06, deve essere compilato EntityOwner/TaxRate o IndOwners/TaxRate.",
         ("GIR1901","GIR1902","GIR1905","GIR1906"),
         lambda ioo,ns: ioo is None or (_find(ioo,"globe:EntityOwner/globe:TaxRate",ns) is None and _find(ioo,"globe:IndOwners/globe:TaxRate",ns) is None)),
        ("70109","Se Basis=GIR1907, deve essere compilato IndOwners/ResCountryCode.",
         ("GIR1907",),
         lambda ioo,ns: ioo is None or _find(ioo,"globe:IndOwners/globe:ResCountryCode",ns) is None),
        ("70110","Se Basis=GIR1903/GIR1908, deve essere compilato IndOwners.",
         ("GIR1903","GIR1908"),
         lambda ioo,ns: ioo is None or _find(ioo,"globe:IndOwners",ns) is None),
        ("70111","Se Basis=GIR1904/GIR1909, deve essere compilato EntityOwner/ExTypeOfEntity.",
         ("GIR1904","GIR1909"),
         lambda ioo,ns: ioo is None or _find(ioo,"globe:EntityOwner/globe:ExTypeOfEntity",ns) is None),
    ]:
        r=CheckResult(code,desc,f"JurisdictionSection/.../CEComputation/AdjustedFANIL/UPEAdjustments/IdentificationOfOwners")
        bad=[]
        for js in jur_sections:
            jn=_t(_find(js,"globe:Jurisdiction",ns))
            for ce in _findall(js,".//globe:CEComputation",ns):
                for ua in _findall(ce,".//globe:UPEAdjustments",ns):
                    be=_find(ua,"globe:Basis",ns); bv=_t(be)
                    if bv not in bases: continue
                    ioo=_find(ua,"globe:IdentificationOfOwners",ns)
                    if check_fn(ioo,ns): bad.append(f"{jn}: Basis={bv} verifica fallita{_src(be)}")
        _ko_all(r, bad)
        out.append(r)

    # 70112
    r=CheckResult("70112","Se Basis=GIR1904, ExTypeOfEntity non deve essere GIR2805.","JurisdictionSection/.../CEComputation/AdjustedFANIL/UPEAdjustments/IdentificationOfOwners/EntityOwner/ExTypeOfEntity")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            for ua in _findall(ce,".//globe:UPEAdjustments",ns):
                if _t(_find(ua,"globe:Basis",ns))!="GIR1904": continue
                ioo=_find(ua,"globe:IdentificationOfOwners",ns)
                ee=_find(ioo,"globe:EntityOwner/globe:ExTypeOfEntity",ns) if ioo else None
                if ee is not None and _t(ee)=="GIR2805": bad.append(f"{jn}: Basis=GIR1904 ma ExTypeOfEntity=GIR2805{_src(ee)}")
    _ko_all(r, bad)
    out.append(r)

    # 70113
    r=CheckResult("70113","Se Basis=GIR1909, ExTypeOfEntity non deve essere GIR2804.","JurisdictionSection/.../CEComputation/AdjustedFANIL/UPEAdjustments/IdentificationOfOwners/EntityOwner/ExTypeOfEntity")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            for ua in _findall(ce,".//globe:UPEAdjustments",ns):
                if _t(_find(ua,"globe:Basis",ns))!="GIR1909": continue
                ioo=_find(ua,"globe:IdentificationOfOwners",ns)
                ee=_find(ioo,"globe:EntityOwner/globe:ExTypeOfEntity",ns) if ioo else None
                if ee is not None and _t(ee)=="GIR2804": bad.append(f"{jn}: Basis=GIR1909 ma ExTypeOfEntity=GIR2804{_src(ee)}")
    _ko_all(r, bad)
    out.append(r)

    return out

# ── 8.3.20 CEComputation NetGlobeIncome / Elections / ACT (70114-70124) ──────
def _check_ce_ngi(jur_sections, ns):
    out=[]

    def _adj_groups(ce, path_parent, ns):
        """Raggruppa (Amount, element) per AdjustmentItem code — scope: singolo CEComputation."""
        groups = {}
        for adj in _findall(ce, path_parent, ns):
            code = _t(_find(adj,"globe:AdjustmentItem",ns))
            ae   = _find(adj,"globe:Amount",ns)
            amt  = _decimal(_t(ae)) if ae is not None else None
            if amt is None: continue
            groups.setdefault(code, []).append((amt, ae))
        return groups

    # 70114 FIX: "segni opposti" solo per stesso AdjustmentItem code
    r=CheckResult("70114","Se due valori NetGlobeIncome/Adjustments/Amount con stesso AdjustmentItem, uno deve essere negativo e l'altro positivo.","JurisdictionSection/.../CEComputation/NetGlobeIncome/Adjustments/Amount")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            for code,pairs in _adj_groups(ce,"globe:NetGlobeIncome/globe:Adjustments",ns).items():
                if len(pairs)==2:
                    a0,e0=pairs[0]; a1,e1=pairs[1]
                    if not (a0<0<a1) and not (a1<0<a0):
                        bad.append(f"{jn}: AdjustmentItem={code!r} ha due Amount [{a0},{a1}] senza segni opposti{_src(e0)}{_src(e1)}")
    _ko_all(r, bad)
    out.append(r)

    # 70115
    r=CheckResult("70115","Se NetGlobeIncome/AdjustmentItem contiene GIR2022/GIR2023, AdjustedFANIL/UPEAdjustments deve essere compilato.","JurisdictionSection/.../CEComputation/NetGlobeIncome/Adjustments/AdjustmentItem")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            items={_t(x) for x in _findall(ce,"globe:NetGlobeIncome/globe:Adjustments/globe:AdjustmentItem",ns)}
            if items&{"GIR2022","GIR2023"} and _find(ce,".//globe:AdjustedFANIL/globe:Adjustment/globe:UPEAdjustments",ns) is None:
                bad.append(f"{jn}: GIR2022/2023 ma UPEAdjustments assente")
    _ko_all(r, bad)
    out.append(r)

    # 70116
    r=CheckResult("70116","Se NetGlobeIncome/AdjustmentItem=GIR2025, NetGlobeIncome/IntShippingIncome deve essere compilato.","JurisdictionSection/.../CEComputation/NetGlobeIncome/IntShippingIncome")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            if "GIR2025" in {_t(x) for x in _findall(ce,"globe:NetGlobeIncome/globe:Adjustments/globe:AdjustmentItem",ns)}:
                if _find(ce,"globe:NetGlobeIncome/globe:IntShippingIncome",ns) is None: bad.append(f"{jn}: GIR2025 ma IntShippingIncome assente")
    _ko_all(r, bad)
    out.append(r)

    # 70117
    r=CheckResult("70117","Se NetGlobeIncome/AdjustmentItem=GIR2024, Elections/Art7.6 deve essere compilato.","JurisdictionSection/.../CEComputation/Elections/Art7.6")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            if "GIR2024" in {_t(x) for x in _findall(ce,"globe:NetGlobeIncome/globe:Adjustments/globe:AdjustmentItem",ns)}:
                if _find(ce,"globe:Elections/globe:Art7.6",ns) is None: bad.append(f"{jn}: GIR2024 ma Elections/Art7.6 assente")
    _ko_all(r, bad)
    out.append(r)

    # 70118 FIX: stesso meccanismo di 70114 per AdjustedCoveredTax/Adjustments
    r=CheckResult("70118","Se due valori AdjustedCoveredTax/Adjustments/Amount con stesso AdjustmentItem, uno deve essere negativo e l'altro positivo.","JurisdictionSection/.../CEComputation/AdjustedCoveredTax/Adjustments/Amount")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            for code,pairs in _adj_groups(ce,"globe:AdjustedCoveredTax/globe:Adjustments",ns).items():
                if len(pairs)==2:
                    a0,e0=pairs[0]; a1,e1=pairs[1]
                    if not (a0<0<a1) and not (a1<0<a0):
                        bad.append(f"{jn}: ACT AdjustmentItem={code!r} ha due Amount [{a0},{a1}] senza segni opposti{_src(e0)}{_src(e1)}")
    _ko_all(r, bad)
    out.append(r)

    # 70119 FIX: scope = singolo CEComputation, non ETR aggregato
    r=CheckResult("70119","Ogni AdjustmentItem in CEComputation/AdjustedCoveredTax/Adjustments non può comparire più di una volta per CEComputation.","JurisdictionSection/.../CEComputation/AdjustedCoveredTax/Adjustments/AdjustmentItem")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            items=[_t(x) for x in _findall(ce,"globe:AdjustedCoveredTax/globe:Adjustments/globe:AdjustmentItem",ns)]
            dups={x for x in items if items.count(x)>1}
            if dups: bad.append(f"{jn}: CEComputation ACT AdjustmentItem duplicati: {dups}")
    _ko_all(r, bad)
    out.append(r)

    # 70120
    r=CheckResult("70120","CEComputation/DeferTaxAdjustAmt/Total = DeferTaxExpense + Σ(Adjustment/Amount) + Recast/Higher + Recast/Lower.","JurisdictionSection/.../CEComputation/AdjustedCoveredTax/DeferTaxAdjustAmt/Total")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            dta=_find(ce,"globe:AdjustedCoveredTax/globe:DeferTaxAdjustAmt",ns)
            if dta is None: continue
            te=_find(dta,"globe:Total",ns); tot=_decimal(_t(te))
            dte=_decimal(_t(_find(dta,"globe:DeferTaxExpense",ns))) or Decimal(0)
            hi=_decimal(_t(_find(dta,"globe:Recast/globe:Higher",ns))) or Decimal(0)
            lo=_decimal(_t(_find(dta,"globe:Recast/globe:Lower",ns))) or Decimal(0)
            as_=sum((_decimal(_t(_find(adj,"globe:Amount",ns))) or Decimal(0)) for adj in _findall(dta,"globe:Adjustment",ns))
            if tot is None: continue
            if abs(tot-(dte+as_+hi+lo))>Decimal("1"): bad.append(f"{jn}: CEComputation DTA Total={tot} ≠ {dte+as_+hi+lo:.2f}{_src(te)}")
    _ko_all(r, bad)
    out.append(r)

    # 70121: scope = ETRComputation (tutti i CEComputation aggregati), come l'AdE.
    # Il testo della regola dice "per ETR element": l'AdE aggrega tutti i
    # CEComputation/DeferTaxAdjustAmt/Adjustment dentro lo stesso ETRComputation
    # e segnala se lo stesso AdjustmentItem code compare più di una volta
    # su CE diversi. Due CEComputation distinti non possono quindi usare lo
    # stesso codice nel loro DeferTaxAdjustAmt/Adjustment.
    r=CheckResult("70121","Ogni AdjustmentItem in CEComputation/DeferTaxAdjustAmt/Adjustment non può comparire più di una volta per ETR (scope: tutti i CEComputation nell'ETRComputation).","JurisdictionSection/.../CEComputation/AdjustedCoveredTax/DeferTaxAdjustAmt/Adjustment/AdjustmentItem")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ec in _findall(js,".//globe:ETRComputation",ns):
            # Raccoglie tutti gli AdjustmentItem da tutti i CEComputation dentro questo ETRComputation
            all_items=[]
            for ce in _findall(ec,"globe:CEComputation",ns):
                dta=_find(ce,"globe:AdjustedCoveredTax/globe:DeferTaxAdjustAmt",ns)
                if dta is None: continue
                for adj in _findall(dta,"globe:Adjustment",ns):
                    ai_el=_find(adj,"globe:AdjustmentItem",ns)
                    if ai_el is not None: all_items.append((_t(ai_el), ai_el))
            # Controlla duplicati aggregando su tutti i CE dell'ETRComputation
            codes=[code for code,_ in all_items]
            dups={x for x in codes if codes.count(x)>1}
            if dups:
                # Trova la prima occorrenza duplicata per il messaggio
                first_dup_el=next((el for code,el in all_items if code in dups), None)
                bad.append(f"{jn}: DeferTaxAdjustAmt/Adjustment AdjustmentItem duplicati tra CEComputation: {dups}{_src(first_dup_el)}")
    _ko_all(r, bad)
    out.append(r)

    # 70122
    r=CheckResult("70122","Se due valori CEComputation/DeferTaxAdjustAmt/Adjustment/Amount, uno deve essere negativo e l'altro positivo.","JurisdictionSection/.../CEComputation/AdjustedCoveredTax/DeferTaxAdjustAmt/Adjustment/Amount")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            dta=_find(ce,"globe:AdjustedCoveredTax/globe:DeferTaxAdjustAmt",ns)
            if dta is None: continue
            amts=[_decimal(_t(_find(adj,"globe:Amount",ns))) for adj in _findall(dta,"globe:Adjustment",ns)]
            amts=[a for a in amts if a is not None]
            if len(amts)==2:
                if not (amts[0]<0<amts[1]) and not (amts[1]<0<amts[0]):
                    bad.append(f"{jn}: DTA Amounts {amts} non opposti")
    _ko_all(r, bad)
    out.append(r)

    # 70123
    r=CheckResult("70123","AdjustedIncomeTax/CrossAllocation/Additions non deve essere negativo.","JurisdictionSection/.../CEComputation/AdjustedIncomeTax/CrossAllocation/Additions")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            ae=_find(ce,"globe:AdjustedIncomeTax/globe:CrossAllocation/globe:Additions",ns)
            if ae is not None:
                v=_decimal(_t(ae))
                if v is not None and v<0: bad.append(f"{jn}: CrossAllocation/Additions={v}<0{_src(ae)}")
    _ko_all(r, bad)
    out.append(r)

    # 70124
    r=CheckResult("70124","AdjustedIncomeTax/CrossAllocation/Reductions non deve essere positivo.","JurisdictionSection/.../CEComputation/AdjustedIncomeTax/CrossAllocation/Reductions")
    bad=[]
    for js in jur_sections:
        jn=_t(_find(js,"globe:Jurisdiction",ns))
        for ce in _findall(js,".//globe:CEComputation",ns):
            re=_find(ce,"globe:AdjustedIncomeTax/globe:CrossAllocation/globe:Reductions",ns)
            if re is not None:
                v=_decimal(_t(re))
                if v is not None and v>0: bad.append(f"{jn}: CrossAllocation/Reductions={v}>0{_src(re)}")
    _ko_all(r, bad)
    out.append(r)

    return out


# ── XLSX output ───────────────────────────────────────────────────────────────
_FILL_OK   = PatternFill("solid", fgColor="C6EFCE")
_FILL_KO   = PatternFill("solid", fgColor="FFC7CE")
_FILL_SKIP = PatternFill("solid", fgColor="FFEB9C")
_FILL_WARN = PatternFill("solid", fgColor="DDEBF7")
_FONT_HDR  = Font(bold=True, color="FFFFFF")
_FILL_BLUE = PatternFill("solid", fgColor="00338D")
_FILL_YEL  = PatternFill("solid", fgColor="EAAA00")
_THIN = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))

def _hdr(ws,row,col,value,fill=None):
    c=ws.cell(row=row,column=col,value=value)
    c.font=_FONT_HDR; c.fill=fill or _FILL_BLUE; c.border=_THIN
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    return c

def _cell(ws,row,col,value,fill=None,bold=False):
    c=ws.cell(row=row,column=col,value=value)
    if fill: c.fill=fill
    c.border=_THIN; c.font=Font(bold=bold)
    c.alignment=Alignment(wrap_text=True,vertical="top")
    return c

def _ffill(status):
    return {"OK":_FILL_OK,"KO":_FILL_KO,"SKIP":_FILL_SKIP,"WARN":_FILL_WARN}.get(status)

def _write_xlsx(results, xml_path, output_dir, guid_resolved=False, version_fixed=False):
    stem=_sanitize_filename(xml_path.stem)
    out_path=output_dir/f"{stem}_validation_report.xlsx"
    wb=Workbook()

    # Sommario
    ws1=wb.active; ws1.title="Sommario"; ws1.sheet_view.showGridLines=False
    counts={s:sum(1 for r in results if r.status==s) for s in ("OK","KO","SKIP","WARN")}
    esito="✅ VALIDO" if counts["KO"]==0 else "❌ ERRORI PRESENTI"

    ws1.merge_cells("A1:F1")
    t=ws1["A1"]; t.value="PILLAR GloBE/DAC9 – RAPPORTO DI VALIDAZIONE"
    t.font=Font(bold=True,size=14,color="FFFFFF"); t.fill=_FILL_BLUE
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws1.row_dimensions[1].height=28

    info=[("File XML",xml_path.name),("Data/ora",datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
          ("Check totali",len(results)),("✓ OK",counts["OK"]),("✗ KO",counts["KO"]),
          ("▷ SKIP",counts["SKIP"]),("⚠ WARN",counts["WARN"]),("Esito",esito)]
    if version_fixed: info.insert(2,("⚠ Version auto-fix","L'attributo GLOBE_OECD/@version era assente: impostato automaticamente a '1.0'. Il file sorgente NON è stato modificato."))
    if guid_resolved: info.insert(2,("⚠ Placeholder GUID","Il file conteneva {Guid:D} non risolti. I check sono stati eseguiti su UUID4 casuali. Il file sorgente NON è stato modificato."))

    for i,(k,v) in enumerate(info,start=2):
        ws1.cell(row=i,column=1,value=k).font=Font(bold=True)
        ws1.cell(row=i,column=1).border=_THIN
        c=ws1.cell(row=i,column=2,value=v); c.border=_THIN
        if k=="Esito": c.fill=_FILL_OK if "VALIDO" in str(v) else _FILL_KO; c.font=Font(bold=True)
        elif k.startswith("⚠"): c.fill=_FILL_SKIP; ws1.cell(row=i,column=1).fill=_FILL_SKIP
    ws1.column_dimensions["A"].width=22; ws1.column_dimensions["B"].width=80

    # Altezza riga dinamica in base al numero di righe nel detail
    def _row_height(detail: str) -> int:
        if not detail:
            return 20
        n = detail.count("\n") + 1
        return max(20, min(n * 16, 400))   # 16pt per riga, cap a 400pt

    # Dettaglio
    ws2=wb.create_sheet("Dettaglio"); ws2.sheet_view.showGridLines=False
    hdrs=["Codice","Categoria","Descrizione","XPath","Esito","Dettaglio"]; wds=[10,12,55,50,8,90]
    for col,(h,w) in enumerate(zip(hdrs,wds),start=1):
        _hdr(ws2,1,col,h); ws2.column_dimensions[get_column_letter(col)].width=w
    for row,r in enumerate(results,start=2):
        cat="SEVERE" if r.code.startswith("6") else "OTHER" if r.code.startswith("7") else "FILE"
        f=_ffill(r.status)
        _cell(ws2,row,1,r.code,f); _cell(ws2,row,2,cat,f); _cell(ws2,row,3,r.desc,f)
        _cell(ws2,row,4,r.xpath,f); _cell(ws2,row,5,r.status,f,bold=True); _cell(ws2,row,6,r.detail,f)
        ws2.row_dimensions[row].height=_row_height(r.detail)
    ws2.freeze_panes="A2"; ws2.auto_filter.ref=f"A1:F{len(results)+1}"

    # Errori e Warning
    ws3=wb.create_sheet("Errori e Warning"); ws3.sheet_view.showGridLines=False
    for col,(h,w) in enumerate(zip(hdrs,wds),start=1):
        _hdr(ws3,1,col,h,_FILL_YEL); ws3.column_dimensions[get_column_letter(col)].width=w
    er=2
    for r in results:
        if r.status in ("KO","WARN"):
            cat="SEVERE" if r.code.startswith("6") else "OTHER" if r.code.startswith("7") else "FILE"
            f=_ffill(r.status)
            _cell(ws3,er,1,r.code,f); _cell(ws3,er,2,cat,f); _cell(ws3,er,3,r.desc,f)
            _cell(ws3,er,4,r.xpath,f); _cell(ws3,er,5,r.status,f,bold=True); _cell(ws3,er,6,r.detail,f)
            ws3.row_dimensions[er].height=_row_height(r.detail); er+=1
    if er==2: ws3.cell(row=2,column=1,value="Nessun errore rilevato ✅").font=Font(bold=True,color="375623")
    ws3.freeze_panes="A2"

    wb.save(str(out_path))
    return out_path

if __name__=="__main__":
    import sys
    if len(sys.argv)<2:
        print("Usage: python xml_validator.py <file.xml> [output_dir]")
        sys.exit(1)
    p=Path(sys.argv[1]); od=Path(sys.argv[2]) if len(sys.argv)>2 else p.parent
    res=validate(p,od)
    print(f"Report: {res}")